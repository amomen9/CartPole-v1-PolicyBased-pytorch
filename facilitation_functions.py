#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Facilitation functions for the experiment pipeline.
Contains file/Excel utility helpers and algorithm job builders.
"""

import os
import shutil
import time
import glob
from copy import copy
from datetime import datetime
from functools import lru_cache
from concurrent.futures import ProcessPoolExecutor, as_completed
from typing import Any, Tuple

import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from functions import (
    LearningCurvePlot,
    smooth,
    _apply_optional_smoothing,
    _load_benchmark_curve,
    average_over_repetitions,
)


# ── File / Excel utilities ────────────────────────────────────────────────────

def _fmt(v):
    """Format numbers safely for filenames."""
    if isinstance(v, float):
        return f"{v:.3g}".replace(".", "p")
    return str(v)


def _format_legend_value(val: Any) -> str:
    """Format legend values, unwrapping single-item containers."""
    if isinstance(val, np.ndarray):
        values = np.atleast_1d(val).tolist()
    elif isinstance(val, (list, tuple, set)):
        values = list(val)
    elif hasattr(val, "__iter__") and not isinstance(val, (str, bytes, dict)):
        values = list(val)
    else:
        return _fmt_legend(val)

    if len(values) == 0:
        return "[]"
    if len(values) == 1:
        return _fmt_legend(values[0])
    return "[" + ",".join(_fmt_legend(item) for item in values) + "]"


def _format_legend_label(label: str) -> str:
    """Return legend labels exactly as configured in Experiment.py."""
    return str(label)


def _excel_cell_display_width(value: Any) -> int:
    """Return the visible width of a value when exported to Excel."""
    if value is None:
        return 0
    if isinstance(value, (float, np.floating)) and np.isnan(value):
        return 0
    text = str(value)
    if not text:
        return 0
    return max(len(line) for line in text.splitlines())


def _autosize_excel_columns(worksheet, dataframe) -> None:
    """Resize Excel columns so their contents fit the widest cell in each column."""
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        max_width = _excel_cell_display_width(column_name)
        for value in dataframe[column_name].tolist():
            max_width = max(max_width, _excel_cell_display_width(value))
        worksheet.column_dimensions[get_column_letter(column_index)].width = max_width


def _set_raw_excel_column_widths(worksheet, headers, first_row_values) -> None:
    """Set raw-sheet widths from the header and first data row only."""
    if first_row_values is None:
        first_row_values = []
    for column_index, header in enumerate(headers, start=1):
        first_value = first_row_values[column_index - 1] if column_index - 1 < len(first_row_values) else ""
        width = max(_excel_cell_display_width(header), _excel_cell_display_width(first_value)) + 1
        worksheet.column_dimensions[get_column_letter(column_index)].width = width


def _sheet_hp_text(value: Any) -> str:
    """Return a stable string representation for Excel hyperparameter values."""
    if value is None:
        return ""
    if isinstance(value, np.ndarray):
        if value.ndim == 0:
            value = value.item()
        else:
            value = value.tolist()
    if isinstance(value, (list, tuple, set)):
        items = list(value)
        if len(items) == 0:
            return "[]"
        if len(items) == 1:
            return _sheet_hp_text(items[0])
        return "[" + ",".join(_sheet_hp_text(item) for item in items) + "]"
    if isinstance(value, np.bool_):
        value = bool(value)
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (np.integer, int)):
        return str(int(value))
    if isinstance(value, (np.floating, float)):
        number = float(value)
        if number.is_integer():
            return str(int(number))
        return format(number, ".6g")
    if isinstance(value, str):
        text = value.strip()
        if text.startswith("[") and text.endswith("]"):
            try:
                import ast
                parsed = ast.literal_eval(text)
                return _sheet_hp_text(parsed)
            except Exception:
                return text.replace(", ", ",")
        return text
    return str(value).strip()


def _sheet_hp_text_candidates(value: Any) -> list[str]:
    """Return the normalized text candidate(s) for a config value."""
    return [_sheet_hp_text(value)]


def _empty_data_sheets_dir(dir_path):
    if os.path.isdir(dir_path):
        shutil.rmtree(dir_path)
    os.makedirs(dir_path, exist_ok=True)


def _save_results_to_excel(dir_path, base_filename, setting_jobs, setting_results):
    import pandas as pd
    filepath = os.path.join(dir_path, f"{base_filename}.xlsx")
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        for idx, (job, result) in enumerate(zip(setting_jobs, setting_results)):
            learning_curve, learning_curve_std, timesteps = result
            sheet_name = f"Setting_{idx + 1:03d}"
            data = {
                "timestep": timesteps,
                "learning_curve_mean": learning_curve,
                "learning_curve_std": learning_curve_std,
            }
            hp = job["hyperparams"]
            for key, value in hp.items():
                data[key] = value
            data["curve_label"] = job["curve_label"]
            df = pd.DataFrame(data)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            _autosize_excel_columns(writer.sheets[sheet_name], df)
    print(f"Saved {len(setting_results)} settings to {filepath}")


def _load_results_from_excel(filepath, algo_config: dict[str, Any] | None, formatted_sheets: bool = False):
    """Load results from an Excel file, validating each sheet's hyperparameters
    against the current algo config.

    Workbook layout assumption:
    - Row 1 is the title / merged heading row.
    - Row 2 contains the actual column headers.
    - Data starts on row 3.

    Returns a tuple (results, mismatches):
        results: list of dicts [{"learning_curve", "learning_curve_std", "timesteps", "curve_label"}, ...]
        mismatches: dict {param_name: (sheet_value, config_value)} for the first mismatch
                    encountered per parameter across all skipped sheets.
    Sheets that don't match are skipped.
    """
    import pandas as pd

    algo_config = algo_config or {}
    header_row = 1 if formatted_sheets else 0
    try:
        sheets = pd.read_excel(filepath, sheet_name=None, engine="openpyxl", header=header_row)
    except Exception as exc:
        raise ValueError(f"Failed to read Excel file '{filepath}': {exc}") from exc

    def _extract_matching_results(sheet_map):
        basename = os.path.basename(filepath)
        algo_prefix = os.path.splitext(basename)[0].upper()
        legend: dict[str, tuple[str, bool]] = _resolve_legend_flags(algo_config or {}, warn_on_suppression=False)
        skip_keys = {"legend_parameters", "nn_include_hp_in_legend", "nn_include_lr_in_legend"}

        extracted_results = []
        extracted_mismatches = {}
        required_columns = {"timestep", "learning_curve_mean", "learning_curve_std"}

        for sheet_name in sorted(sheet_map.keys()):
            df = sheet_map[sheet_name]
            if not required_columns.issubset(set(df.columns)):
                continue

            sheet_matched = True
            if algo_config:
                for cfg_key, cfg_val in algo_config.items():
                    if cfg_key in skip_keys:
                        continue
                    if cfg_key not in df.columns:
                        # Only invalidate reuse when the current run depends on these eval-mode keys.
                        if cfg_key == "eval_with_env_episode_trials":
                            if bool(cfg_val) is True:
                                sheet_matched = False
                                if cfg_key not in extracted_mismatches:
                                    extracted_mismatches[cfg_key] = ("<missing_column>", cfg_val)
                        elif cfg_key == "n_eval_episodes":
                            if bool(algo_config.get("eval_with_env_episode_trials", False)) is True:
                                sheet_matched = False
                                if cfg_key not in extracted_mismatches:
                                    extracted_mismatches[cfg_key] = ("<missing_column>", cfg_val)
                        continue

                    sheet_val = df[cfg_key].iloc[0]
                    sheet_val_text = _sheet_hp_text(_parse_sheet_value(sheet_val))
                    cfg_val_candidates = _sheet_hp_text_candidates(cfg_val)

                    if sheet_val_text not in cfg_val_candidates:
                        sheet_matched = False
                        if cfg_key not in extracted_mismatches:
                            extracted_mismatches[cfg_key] = (sheet_val_text, cfg_val)

            if not sheet_matched:
                continue

            try:
                timesteps = df["timestep"].values.astype(np.int32)
                learning_curve = df["learning_curve_mean"].values.astype(np.float32)
                learning_curve_std = df["learning_curve_std"].values.astype(np.float32)
            except Exception:
                continue

            label_parts = [algo_prefix]
            for legend_key, (legend_label, show) in legend.items():
                if not show:
                    continue
                if legend_key in df.columns:
                    val = _parse_sheet_value(df[legend_key].iloc[0])
                elif legend_key in algo_config:
                    val = algo_config[legend_key]
                else:
                    continue

                legend_label = _format_legend_label(legend_label)
                if isinstance(val, bool):
                    label_parts.append(f"{legend_label}{val}")
                else:
                    label_parts.append(f"{legend_label}{_format_legend_value(val)}")

            extracted_results.append({
                "learning_curve": learning_curve,
                "learning_curve_std": learning_curve_std,
                "timesteps": timesteps,
                "curve_label": ", ".join(label_parts),
            })

        return extracted_results, extracted_mismatches

    results, mismatches = _extract_matching_results(sheets)

    # Fallback for any older workbook layout that uses the opposite header row.
    if not results:
        try:
            fallback_header = 0 if formatted_sheets else 1
            fallback_sheets = pd.read_excel(filepath, sheet_name=None, engine="openpyxl", header=fallback_header)
        except Exception:
            return results, mismatches
        fallback_results, fallback_mismatches = _extract_matching_results(fallback_sheets)
        if fallback_results:
            return fallback_results, fallback_mismatches

    return results, mismatches


def _parse_sheet_value(val):
    """Parse a value read from an Excel HP column into a stable string."""
    return _sheet_hp_text(val)


def _hp_value_matches(cfg_key, cfg_val, sheet_val):
    """Check whether a sheet's HP value is compatible with the config value."""
    sheet_text = _sheet_hp_text(sheet_val)
    return sheet_text in _sheet_hp_text_candidates(cfg_val)


def _values_equal(a, b):
    """Compare two scalar values with float tolerance."""
    try:
        fa, fb = float(a), float(b)
        return abs(fa - fb) < 1e-7
    except (TypeError, ValueError):
        pass
    # Handle booleans explicitly
    if isinstance(a, (bool, np.bool_)) or isinstance(b, (bool, np.bool_)):
        return bool(a) == bool(b)
    return str(a) == str(b)


def _load_all_excel_curves(data_sheets_dir, algo_configs=None, formatted_sheets: bool = False):
    """Load all .xlsx files from data_sheets_dir (non-recursive).

    algo_configs: dict mapping algo name (e.g. "REINFORCE", "DQN") to its config dict.
    Each file's algo is inferred from the filename stem, so only files named
    like ``REINFORCE.xlsx`` or ``DQN.xlsx`` are matched.
    Sheets are filtered by HP value matching and labels are built using legend_parameters,
    both delegated to _load_results_from_excel.

    Returns a list of dicts: [{curve_label, learning_curve, learning_curve_std, timesteps, source_file}, ...]
    """
    all_curves = []
    pattern = os.path.join(data_sheets_dir, "*.xlsx")
    for filepath in sorted(glob.glob(pattern)):
        basename = os.path.basename(filepath)
        algo_prefix = os.path.splitext(basename)[0].upper()
        algo_config = (algo_configs or {}).get(algo_prefix)
        try:
            results, _ = _load_results_from_excel(filepath, algo_config, formatted_sheets=formatted_sheets)
        except Exception:
            continue
        for entry in results:
            entry["source_file"] = basename
            all_curves.append(entry)
    return all_curves


# ── Per-algorithm filename builder ────────────────────────────────────────────

def _build_algo_filename(algo_name, algo_config, n_repetitions, n_timesteps, eval_interval):
    """Build the workbook filename stem for a single algorithm."""
    return build_algorithm_filename(algo_name)


# ── Legend parameter helpers ──────────────────────────────────────────────────

LegendEntry = Tuple[str, bool]


def _normalize_legend_entry(param_name: str, entry: Any) -> LegendEntry:
    """Normalize a legend entry to a (label, show_flag) pair."""
    if isinstance(entry, (list, tuple)) and len(entry) == 2:
        label, show = entry
    else:
        label, show = param_name, entry
    return str(label), bool(show)


def _resolve_legend_flags(cfg: dict[str, Any], *, warn_on_suppression: bool = True) -> dict[str, LegendEntry]:
    """Resolve which parameters should appear in the legend.

    Uses ``legend_parameters`` dict if present. Falls back to (and is
    overridden by) the legacy ``nn_include_hp_in_legend`` /
    ``nn_include_lr_in_legend`` flags for backward compatibility.

    When DQN epsilon decay is disabled (``epsilon_decay_interval == 0``),
    suppress the dependent epsilon-decay fields from the legend so the
    DQN labels match the shared REINFORCE/AC/A2C path. If
    ``warn_on_suppression`` is True, print one warning per suppressed field.

    Returns a dict {param_name: (display_label, bool)}.
    """
    legend: dict[str, LegendEntry] = {}
    raw_legend_parameters = cfg.get("legend_parameters", {})
    if isinstance(raw_legend_parameters, dict):
        for param_name, entry in raw_legend_parameters.items():
            legend[param_name] = _normalize_legend_entry(param_name, entry)

    # Legacy overrides
    if "nn_include_lr_in_legend" in cfg:
        lr_key = "learning_rate" if "learning_rate" in cfg else "actor_lr"
        current_lr_entry = legend.get(lr_key)
        if current_lr_entry is not None:
            current_lr_label = current_lr_entry[0]
        else:
            current_lr_label = lr_key
        legend[lr_key] = (current_lr_label, bool(cfg["nn_include_lr_in_legend"]))
    if "nn_include_hp_in_legend" in cfg:
        hp_flag = bool(cfg["nn_include_hp_in_legend"])
        nn_key = "nn_hidden_layer_widths" if "nn_hidden_layer_widths" in cfg else "actor_hidden_nn"
        current_nn_entry = legend.get(nn_key)
        if current_nn_entry is not None:
            current_nn_label = current_nn_entry[0]
        else:
            current_nn_label = nn_key
        legend[nn_key] = (current_nn_label, hp_flag)
        current_gamma_entry = legend.get("gamma")
        if current_gamma_entry is not None:
            current_gamma_label = current_gamma_entry[0]
        else:
            current_gamma_label = "gamma"
        legend["gamma"] = (current_gamma_label, hp_flag)

    if cfg.get("exploration_method", "egreedy") == "egreedy":
        try:
            epsilon_decay_interval = int(cfg.get("epsilon_decay_interval", 1))
        except (TypeError, ValueError):
            epsilon_decay_interval = 1
        if epsilon_decay_interval == 0:
            for key in ("epsilon_start", "epsilon_end", "epsilon_decay"):
                entry = legend.get(key)
                if entry is None or not entry[1]:
                    continue
                if warn_on_suppression:
                    display_label = entry[0] if str(entry[0]).strip() else key
                    print(
                        f"[DQN] Warning (plot): '{display_label}' legend entry suppressed because "
                        "epsilon_decay_interval=0 disables epsilon decay trials."
                    )
                legend[key] = (entry[0], False)

    return legend


def _fmt_legend(v: Any) -> str:
    """Format numbers for legend display (keeps decimal point)."""
    if isinstance(v, float):
        return f"{v:.3g}"
    return str(v)


def _build_legend_parts(legend: dict[str, LegendEntry], cfg: dict[str, Any]) -> list[str]:
    """Build a list of 'label=value' strings for every legend-enabled parameter."""
    parts: list[str] = []
    for key, (label, show) in legend.items():
        if not show:
            continue
        val = cfg.get(key)
        if val is None:
            continue
        label = _format_legend_label(label)
        if isinstance(val, bool):
            parts.append(f"{label}{val}")
        else:
            parts.append(f"{label}{_format_legend_value(val)}")
    return parts


# ── Build setting jobs per algorithm ──────────────────────────────────────────

def _parse_pg_config(cfg):
    """Parse policy-gradient config into sweepable arrays (with backward compat)."""
    gammas = np.atleast_1d(np.asarray(cfg.get("gamma", np.array([0.99])), dtype=np.float32))
    learning_rates = np.atleast_1d(np.asarray(cfg.get("actor_lr", np.array([0.001])), dtype=np.float32))

    raw_nn = cfg.get("actor_hidden_nn", [[32, 32]])
    if isinstance(raw_nn, np.ndarray) and raw_nn.ndim == 1:
        nn_architectures = [raw_nn]  # backward compat: np.array([32,32])
    elif isinstance(raw_nn, list) and len(raw_nn) > 0 and not isinstance(raw_nn[0], (list, np.ndarray)):
        nn_architectures = [np.asarray(raw_nn, dtype=np.int32)]  # backward compat: flat list [32,32]
    else:
        nn_architectures = [np.asarray(arch, dtype=np.int32) for arch in raw_nn]

    legend = _resolve_legend_flags(cfg)
    return gammas, learning_rates, nn_architectures, legend


def _build_reinforce_jobs(
    *,
    algo_config,
    n_repetitions,
    n_timesteps,
    eval_interval,
    max_train_episode_length,
    max_eval_episode_length,
    base_seed,
    eval_with_env_episode_trials: bool,
    n_eval_episodes: int,
):
    cfg = algo_config
    gammas, learning_rates, nn_architectures, legend = _parse_pg_config(cfg)

    setting_jobs = []
    for gamma_val in gammas:
        gamma_val = float(gamma_val)
        for nn_arch in nn_architectures:
            nn_arch = np.asarray(nn_arch, dtype=np.int32)
            for lr_val in learning_rates:
                lr_val = float(lr_val)
                iter_cfg = {**cfg, "gamma": gamma_val, "actor_lr": lr_val, "actor_hidden_nn": nn_arch}
                label_parts = ["REINFORCE"] + _build_legend_parts(legend, iter_cfg)
                curve_label = ", ".join(label_parts)
                setting_jobs.append({
                    "curve_label": curve_label,
                    "method": "REINFORCE",
                    "kwargs": dict(
                        method="REINFORCE",
                        n_repetitions=n_repetitions,
                        n_timesteps=n_timesteps,
                        eval_interval=eval_interval,
                        max_train_episode_length=max_train_episode_length,
                        max_eval_episode_length=max_eval_episode_length,
                        actor_lr=lr_val,
                        gamma=gamma_val,
                        actor_hidden_nn=nn_arch,
                        base_seed=base_seed,
                        eval_with_env_episode_trials=eval_with_env_episode_trials,
                        n_eval_episodes=n_eval_episodes,
                        plot_smoothing_window=1,
                    ),
                    "hyperparams": {
                        "n_repetitions": n_repetitions,
                        "n_timesteps": n_timesteps,
                        "eval_interval": eval_interval,
                        "max_train_episode_length": max_train_episode_length,
                        "max_eval_episode_length": max_eval_episode_length,
                        "actor_lr": lr_val,
                        "gamma": gamma_val,
                        "actor_hidden_nn": str(nn_arch.tolist()),
                        "eval_with_env_episode_trials": eval_with_env_episode_trials,
                        "n_eval_episodes": n_eval_episodes,
                    },
                })
    return setting_jobs


def _build_ac_jobs(
    *,
    algo_config,
    n_repetitions,
    n_timesteps,
    eval_interval,
    max_train_episode_length,
    max_eval_episode_length,
    base_seed,
    eval_with_env_episode_trials: bool,
    n_eval_episodes: int,
):
    cfg = algo_config
    gammas, learning_rates, nn_architectures, legend = _parse_pg_config(cfg)
    critic_learning_rates = np.atleast_1d(np.asarray(cfg.get("critic_lr", np.array([0.001])), dtype=np.float32))
    raw_critic_nn = cfg.get("critic_hidden_nn", [[64, 64]])
    if isinstance(raw_critic_nn, np.ndarray) and raw_critic_nn.ndim == 1:
        critic_architectures = [raw_critic_nn]
    elif isinstance(raw_critic_nn, list) and len(raw_critic_nn) > 0 and not isinstance(raw_critic_nn[0], (list, np.ndarray)):
        critic_architectures = [np.asarray(raw_critic_nn, dtype=np.int32)]
    else:
        critic_architectures = [np.asarray(arch, dtype=np.int32) for arch in raw_critic_nn]

    setting_jobs = []
    for gamma_val in gammas:
        gamma_val = float(gamma_val)
        for actor_nn in nn_architectures:
            actor_nn = np.asarray(actor_nn, dtype=np.int32)
            for actor_lr_val in learning_rates:
                actor_lr_val = float(actor_lr_val)
                for critic_nn in critic_architectures:
                    critic_nn = np.asarray(critic_nn, dtype=np.int32)
                    for critic_lr_val in critic_learning_rates:
                        critic_lr_val = float(critic_lr_val)
                        iter_cfg = {
                            **cfg,
                            "gamma": gamma_val,
                            "actor_lr": actor_lr_val,
                            "actor_hidden_nn": actor_nn,
                            "critic_lr": critic_lr_val,
                            "critic_hidden_nn": critic_nn,
                        }
                        label_parts = ["AC"] + _build_legend_parts(legend, iter_cfg)
                        curve_label = ", ".join(label_parts)
                        setting_jobs.append({
                            "curve_label": curve_label,
                            "method": "ac",
                            "kwargs": dict(
                                method="ac",
                                n_repetitions=n_repetitions,
                                n_timesteps=n_timesteps,
                                eval_interval=eval_interval,
                                max_train_episode_length=max_train_episode_length,
                                max_eval_episode_length=max_eval_episode_length,
                                actor_lr=actor_lr_val,
                                critic_lr=critic_lr_val,
                                gamma=gamma_val,
                                actor_hidden_nn=actor_nn,
                                critic_hidden_nn=critic_nn,
                                base_seed=base_seed,
                                eval_with_env_episode_trials=eval_with_env_episode_trials,
                                n_eval_episodes=n_eval_episodes,
                                plot_smoothing_window=1,
                            ),
                            "hyperparams": {
                                "n_repetitions": n_repetitions,
                                "n_timesteps": n_timesteps,
                                "eval_interval": eval_interval,
                                "max_train_episode_length": max_train_episode_length,
                                "max_eval_episode_length": max_eval_episode_length,
                                "actor_lr": actor_lr_val,
                                "critic_lr": critic_lr_val,
                                "gamma": gamma_val,
                                "actor_hidden_nn": str(actor_nn.tolist()),
                                "critic_hidden_nn": str(critic_nn.tolist()),
                                "eval_with_env_episode_trials": eval_with_env_episode_trials,
                                "n_eval_episodes": n_eval_episodes,
                            },
                        })
    return setting_jobs


def _build_a2c_jobs(
    *,
    algo_config,
    n_repetitions,
    n_timesteps,
    eval_interval,
    max_train_episode_length,
    max_eval_episode_length,
    base_seed,
    eval_with_env_episode_trials: bool,
    n_eval_episodes: int,
):
    cfg = algo_config
    gammas, actor_learning_rates, actor_architectures, legend = _parse_pg_config(cfg)
    critic_learning_rates = np.atleast_1d(np.asarray(cfg.get("critic_lr", np.array([0.001])), dtype=np.float32))
    raw_critic_nn = cfg.get("critic_hidden_nn", [[64, 64]])
    if isinstance(raw_critic_nn, np.ndarray) and raw_critic_nn.ndim == 1:
        critic_architectures = [raw_critic_nn]
    elif isinstance(raw_critic_nn, list) and len(raw_critic_nn) > 0 and not isinstance(raw_critic_nn[0], (list, np.ndarray)):
        critic_architectures = [np.asarray(raw_critic_nn, dtype=np.int32)]
    else:
        critic_architectures = [np.asarray(arch, dtype=np.int32) for arch in raw_critic_nn]
    TN_steps = np.atleast_1d(np.asarray(cfg.get("TN_step", np.array([10])), dtype=np.int32))

    setting_jobs = []
    for gamma_val in gammas:
        gamma_val = float(gamma_val)
        for actor_nn in actor_architectures:
            actor_nn = np.asarray(actor_nn, dtype=np.int32)
            for actor_lr_val in actor_learning_rates:
                actor_lr_val = float(actor_lr_val)
                for critic_nn in critic_architectures:
                    critic_nn = np.asarray(critic_nn, dtype=np.int32)
                    for critic_lr_val in critic_learning_rates:
                        critic_lr_val = float(critic_lr_val)
                        for tn_step in TN_steps:
                            tn_step = int(tn_step)
                            iter_cfg = {
                                **cfg,
                                "gamma": gamma_val,
                                "actor_lr": actor_lr_val,
                                "actor_hidden_nn": actor_nn,
                                "critic_lr": critic_lr_val,
                                "critic_hidden_nn": critic_nn,
                                "TN_step": tn_step,
                            }
                            label_parts = ["A2C"] + _build_legend_parts(legend, iter_cfg)
                            curve_label = ", ".join(label_parts)
                            setting_jobs.append({
                                "curve_label": curve_label,
                                "method": "a2c",
                                "kwargs": dict(
                                    method="a2c",
                                    n_repetitions=n_repetitions,
                                    n_timesteps=n_timesteps,
                                    eval_interval=eval_interval,
                                    max_train_episode_length=max_train_episode_length,
                                    max_eval_episode_length=max_eval_episode_length,
                                    actor_lr=actor_lr_val,
                                    critic_lr=critic_lr_val,
                                    gamma=gamma_val,
                                    actor_hidden_nn=actor_nn,
                                    critic_hidden_nn=critic_nn,
                                    TN_step=tn_step,
                                    base_seed=base_seed,
                                    eval_with_env_episode_trials=eval_with_env_episode_trials,
                                    n_eval_episodes=n_eval_episodes,
                                    plot_smoothing_window=1,
                                ),
                                "hyperparams": {
                                    "n_repetitions": n_repetitions,
                                    "n_timesteps": n_timesteps,
                                    "eval_interval": eval_interval,
                                    "max_train_episode_length": max_train_episode_length,
                                    "max_eval_episode_length": max_eval_episode_length,
                                    "actor_lr": actor_lr_val,
                                    "critic_lr": critic_lr_val,
                                    "gamma": gamma_val,
                                    "actor_hidden_nn": str(actor_nn.tolist()),
                                    "critic_hidden_nn": str(critic_nn.tolist()),
                                    "TN_step": tn_step,
                                    "eval_with_env_episode_trials": eval_with_env_episode_trials,
                                    "n_eval_episodes": n_eval_episodes,
                                },
                            })
    return setting_jobs


def _build_dqn_jobs(*, dqn_config, n_repetitions, n_timesteps, eval_interval,
                    max_train_episode_length, max_eval_episode_length, base_seed):
    """Build setting_jobs for DQN using assignment2_repo infrastructure."""
    cfg = dqn_config
    learning_rates = np.atleast_1d(np.asarray(cfg.get("learning_rate", np.array([0.001])), dtype=np.float32))
    raw_nn = cfg.get("nn_hidden_layer_widths", [[128, 128]])
    if isinstance(raw_nn, np.ndarray) and raw_nn.ndim == 1:
        nn_architectures = [raw_nn]
    elif isinstance(raw_nn, list) and len(raw_nn) > 0 and not isinstance(raw_nn[0], (list, np.ndarray)):
        nn_architectures = [np.asarray(raw_nn, dtype=np.int32)]
    else:
        nn_architectures = [np.asarray(arch, dtype=np.int32) for arch in raw_nn]
    gamma = float(cfg.get("gamma", 0.99))
    exploration_method = cfg.get("exploration_method", "egreedy")
    epsilons = np.atleast_1d(np.asarray(cfg.get("epsilons", np.array([0.05])), dtype=np.float32))
    epsilon_start = float(cfg.get("epsilon_start", 1.0))
    epsilon_end = float(cfg.get("epsilon_end", 0.02))
    epsilon_decay = float(cfg.get("epsilon_decay", 0.9995))
    epsilon_decay_interval = int(cfg.get("epsilon_decay_interval", 0))
    softmax_temps = np.atleast_1d(np.asarray(cfg.get("softmax_temps", np.array([1.0, 0.5, 0.1])), dtype=np.float32))

    TN_active = np.atleast_1d(np.asarray(cfg.get("TN_active", np.array([True]))))
    TN_step = np.atleast_1d(np.asarray(cfg.get("TN_step", np.array([100])), dtype=np.int32))
    ER_active = np.atleast_1d(np.asarray(cfg.get("ER_active", np.array([True]))))
    ER_replay_buffer_size = int(cfg.get("ER_replay_buffer_size", 80000))
    ER_batch_size = np.atleast_1d(np.asarray(cfg.get("ER_batch_size", np.array([64])), dtype=np.int32))
    ER_min_replay_size = int(cfg.get("ER_min_replay_size", 2000))
    ER_sample_train_frequency = np.atleast_1d(np.asarray(cfg.get("ER_sample_train_frequency", np.array([1])), dtype=np.int32))
    ER_replay_ratio = float(cfg.get("ER_replay_ratio", 1.0))

    legend = _resolve_legend_flags(cfg)

    epsilon_decay_enabled = epsilon_decay_interval > 0
    skip_decay_trials = exploration_method == "egreedy" and not epsilon_decay_enabled

    eval_with_env_episode_trials = bool(cfg.get("eval_with_env_episode_trials", False))
    n_eval_episodes = int(cfg.get("n_eval_episodes", 5))

    setting_jobs = []
    for nn_arch in nn_architectures:
        nn_arch = np.asarray(nn_arch, dtype=np.int32)
        for tn_active in TN_active:
            tn_active_bool = bool(tn_active)
            effective_tn_steps = TN_step if tn_active_bool else np.array([1])
            for er_active_val in ER_active:
                er_active_bool = bool(er_active_val)
                for er_bs in ER_batch_size:
                    er_bs = int(er_bs)
                    for er_freq in ER_sample_train_frequency:
                        er_freq = int(er_freq)
                        for lr_val in learning_rates:
                            lr_val = float(lr_val)
                            for tn_step in effective_tn_steps:
                                tn_step = int(tn_step)
                                iter_cfg = {**cfg, "nn_hidden_layer_widths": nn_arch,
                                            "learning_rate": lr_val, "gamma": gamma}
                                label_parts = ["DQN"]
                                label_parts.extend(_build_legend_parts(legend, iter_cfg))
                                run_label_prefix = ", ".join(label_parts)

                                er_kwargs = dict(
                                    er_active=er_active_bool,
                                    er_replay_buffer_size=ER_replay_buffer_size,
                                    er_batch_size=er_bs,
                                    er_min_replay_size=ER_min_replay_size,
                                    er_sample_train_frequency=er_freq,
                                    er_replay_ratio=ER_replay_ratio,
                                )
                                common = dict(
                                    n_timesteps=n_timesteps,
                                    max_episode_length=max_train_episode_length,
                                    max_eval_episode_length=max_eval_episode_length,
                                    eval_with_env_episode_trials=eval_with_env_episode_trials,
                                    n_eval_episodes=n_eval_episodes,
                                    learning_rate=lr_val,
                                    gamma=gamma,
                                    nn_hidden_layer_widths=nn_arch,
                                    TN_step=tn_step,
                                    target_network_active=tn_active_bool,
                                    eval_interval=eval_interval,
                                    base_seed=base_seed,
                                    exploration_method=exploration_method,
                                    **er_kwargs,
                                )
                                base_hp = {
                                    "n_repetitions": n_repetitions,
                                    "n_timesteps": n_timesteps,
                                    "eval_interval": eval_interval,
                                    "max_episode_length": max_train_episode_length,
                                    "max_eval_episode_length": max_eval_episode_length,
                                    "eval_with_env_episode_trials": eval_with_env_episode_trials,
                                    "n_eval_episodes": n_eval_episodes,
                                    "learning_rate": lr_val,
                                    "gamma": gamma,
                                    "nn_hidden_layer_widths": str(nn_arch.tolist()),
                                    "exploration_method": exploration_method,
                                    "TN_active": tn_active_bool,
                                    "TN_step": tn_step,
                                    "ER_active": er_active_bool,
                                    "ER_batch_size": er_bs,
                                    "ER_sample_train_frequency": er_freq,
                                }

                                # epsilon-decay trial
                                if not skip_decay_trials and exploration_method == "egreedy":
                                    setting_jobs.append({
                                        "curve_label": run_label_prefix,
                                        "method": "dqn",
                                        "kwargs": {
                                            **common,
                                            "n_repetitions": n_repetitions,
                                            "epsilon_start": epsilon_start,
                                            "epsilon_end": epsilon_end,
                                            "epsilon_decay": epsilon_decay,
                                            "epsilon_decay_interval": epsilon_decay_interval,
                                            "softmax_temp": None,
                                        },
                                        "hyperparams": {**base_hp,
                                            "epsilon_start": epsilon_start,
                                            "epsilon_end": epsilon_end,
                                        },
                                    })

                                # fixed-epsilon trials
                                if exploration_method == "egreedy":
                                    for eps in epsilons:
                                        setting_jobs.append({
                                            "curve_label": run_label_prefix,
                                            "method": "dqn",
                                            "kwargs": {
                                                **common,
                                                "n_repetitions": n_repetitions,
                                                "epsilon_start": float(eps),
                                                "epsilon_end": float(eps),
                                                "epsilon_decay": 1.0,
                                                "epsilon_decay_interval": 1,
                                                "softmax_temp": None,
                                            },
                                            "hyperparams": {**base_hp,
                                                "epsilon_start": float(eps),
                                                "epsilon_end": float(eps),
                                            },
                                        })
                                else:
                                    for temp in softmax_temps:
                                        setting_jobs.append({
                                            "curve_label": run_label_prefix,
                                            "method": "dqn",
                                            "kwargs": {
                                                **common,
                                                "n_repetitions": n_repetitions,
                                                "epsilon_start": None,
                                                "epsilon_end": None,
                                                "epsilon_decay": 1.0,
                                                "epsilon_decay_interval": 1,
                                                "softmax_temp": float(temp),
                                            },
                                            "hyperparams": {**base_hp,
                                                "softmax_temp": float(temp),
                                            },
                                        })
    return setting_jobs


# ── Cross-setting parallel execution engine ───────────────────────────────────────────

def _dqn_job_kwargs_to_trial_common(job_kwargs: dict[str, Any], eval_interval: int) -> dict[str, Any]:
    """Map a DQN job's kwargs dict to the run_dqn_trial_returns keyword arguments."""
    kw: dict[str, Any] = dict(job_kwargs)
    exploration_method = kw.get("exploration_method", "egreedy")
    trial_common: dict[str, Any] = dict(
        n_env_steps=int(kw["n_timesteps"]),
        max_episode_length=int(kw["max_episode_length"]),
        max_eval_episode_length=int(kw.get("max_eval_episode_length", kw["max_episode_length"])),
        eval_with_env_episode_trials=bool(kw.get("eval_with_env_episode_trials", False)),
        n_eval_episodes=int(kw.get("n_eval_episodes", 5)),
        learning_rate=float(kw["learning_rate"]),
        nn_hidden_layer_widths=np.asarray(kw["nn_hidden_layer_widths"], dtype=np.int32),
        discount_factor=float(kw["gamma"]),
        target_network_step=int(kw["TN_step"]),
        target_network_active=bool(kw["target_network_active"]),
        n_returns_interval=eval_interval,
        exploration_method="epsilon_greedy" if exploration_method == "egreedy" else exploration_method,
        epsilon_start=kw.get("epsilon_start"),
        epsilon_end=kw.get("epsilon_end"),
        epsilon_decay=float(kw.get("epsilon_decay", 1.0)),
        epsilon_decay_interval=int(kw.get("epsilon_decay_interval", 1)),
        softmax_temp=kw.get("softmax_temp"),
        er_active=bool(kw.get("er_active", False)),
        er_replay_buffer_size=int(kw.get("er_replay_buffer_size", 10000)),
        er_batch_size=int(kw.get("er_batch_size", 64)),
        er_min_replay_size=int(kw.get("er_min_replay_size", 100)),
        er_sample_train_frequency=int(kw.get("er_sample_train_frequency", 1)),
        er_replay_ratio=float(kw.get("er_replay_ratio", 1.0)),
        emit_trial_header=False,
    )
    return trial_common


def _run_dqn_one_rep(trial_common: dict[str, Any], run_seed: int, rep_index: int, n_repetitions: int):
    """Run one DQN repetition. Pickle-safe for ProcessPoolExecutor."""
    from assignment2_repo.DQN import run_dqn_trial_returns
    return run_dqn_trial_returns(
        seed=run_seed,
        trial_run_index=rep_index + 1,
        total_trial_runs=n_repetitions,
        enable_progress_bar=False,
        **trial_common,
    )


def _run_pending_parallel(pending_settings, n_repetitions, n_timesteps, eval_interval,
                          max_train_episode_length, max_eval_episode_length, base_seed,
                          setting_results: list[tuple[np.ndarray, np.ndarray, np.ndarray] | None]):
    """Run all pending (setting × rep) tasks in a single flat ProcessPoolExecutor.

    pending_settings: list of (global_idx, job)
    Fills setting_results[global_idx] = (lc_mean, lc_std, timesteps) for each entry.
    Progress bars mirror the current per-rep format; a (S{n}) suffix is added when
    more than one setting is running simultaneously.
    """
    from multiprocessing import Manager
    from concurrent.futures import ProcessPoolExecutor
    from tqdm import tqdm
    from functions import _run_single_repetition
    import time as _time

    total_tasks = len(pending_settings) * n_repetitions
    multi = len(pending_settings) > 1

    with Manager() as mgr:
        step_counters = {}
        for sp, (_, job) in enumerate(pending_settings):
            if job["method"] != "dqn":
                for r in range(n_repetitions):
                    step_counters[(sp, r)] = mgr.Value('i', 0)

        max_workers = min(total_tasks, os.cpu_count() or 1)
        with ProcessPoolExecutor(max_workers=max_workers) as executor:
            futures = {}
            for sp, (_, job) in enumerate(pending_settings):
                method = job["method"]
                kw: dict[str, Any] = dict(job["kwargs"])
                for r in range(n_repetitions):
                    run_seed = base_seed + r
                    if method == "dqn":
                        trial_common = _dqn_job_kwargs_to_trial_common(kw, eval_interval)
                        future = executor.submit(
                            _run_dqn_one_rep,
                            trial_common=trial_common,
                            run_seed=run_seed,
                            rep_index=r,
                            n_repetitions=n_repetitions,
                        )
                    else:
                        future = executor.submit(
                            _run_single_repetition,
                            method=method,
                            actor_hidden_nn=kw["actor_hidden_nn"],
                            critic_hidden_nn=kw.get("critic_hidden_nn", np.array([64, 64])),
                            actor_lr=kw["actor_lr"],
                            critic_lr=kw.get("critic_lr", 0.001),
                            gamma=kw["gamma"],
                            max_train_episode_length=max_train_episode_length,
                            max_eval_episode_length=max_eval_episode_length,
                            n_timesteps=n_timesteps,
                            eval_interval=eval_interval,
                            run_seed=run_seed,
                            rep_index=r,
                            n_repetitions=n_repetitions,
                            enable_progress_bar=False,
                            shared_step_counter=step_counters[(sp, r)],
                            eval_with_env_episode_trials=bool(kw.get("eval_with_env_episode_trials", False)),
                            n_eval_episodes=int(kw.get("n_eval_episodes", 5)),
                        )
                    futures[future] = (sp, r)

            pbars = {}
            for sp, (_, job) in enumerate(pending_settings):
                m = job["method"].upper()
                for r in range(n_repetitions):
                    suffix = f" (S{sp + 1})" if multi else ""
                    pbars[(sp, r)] = tqdm(
                        total=n_timesteps,
                        desc=f"{m} Rep {r + 1}/{n_repetitions}{suffix}",
                        unit="step",
                        position=sp * n_repetitions + r,
                        leave=True,
                        dynamic_ncols=True,
                        bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]",
                    )

            rep_results = {}
            done = set()
            try:
                while len(done) < total_tasks:
                    for (sp, r), sc in step_counters.items():
                        cur = sc.value
                        pb = pbars[(sp, r)]
                        delta = cur - pb.n
                        if delta > 0:
                            pb.update(delta)
                    for f in list(futures):
                        if f not in done and f.done():
                            done.add(f)
                            sp, r = futures[f]
                            rep_results[(sp, r)] = f.result()
                            rem = n_timesteps - pbars[(sp, r)].n
                            if rem > 0:
                                pbars[(sp, r)].update(rem)
                    _time.sleep(0.25)
            finally:
                for pb in pbars.values():
                    pb.close()
                print()

    for sp, (global_idx, _) in enumerate(pending_settings):
        returns_list = [np.asarray(rep_results[(sp, r)][0], dtype=np.float32)
                        for r in range(n_repetitions)]
        ts = np.asarray(rep_results[(sp, 0)][1], dtype=np.int32)
        all_ret = np.array(returns_list)
        lc_mean = np.mean(all_ret, axis=0)
        lc_std = (np.std(all_ret, axis=0, ddof=1) if len(returns_list) > 1
                  else np.zeros_like(lc_mean))
        setting_results[global_idx] = (lc_mean, lc_std, ts)


# ── DQN repetitions via assignment2_repo ──────────────────────────────────────────────

def _run_dqn_repetitions(job_kwargs: dict[str, Any]):
    """Run n_repetitions of DQN via assignment2_repo and return (mean, std, timesteps).
    Uses ProcessPoolExecutor for parallel trials, matching assignment2_repo pattern."""
    from assignment2_repo.DQN import run_dqn_trial_returns

    kw: dict[str, Any] = dict(job_kwargs)
    n_repetitions = int(kw.pop("n_repetitions"))
    base_seed = int(kw.pop("base_seed", 42))
    n_timesteps = int(kw["n_timesteps"])
    eval_interval = int(kw.pop("eval_interval"))
    exploration_method = kw.pop("exploration_method")

    # Map field names to assignment2_repo DQN API
    trial_common: dict[str, Any] = dict(
        n_env_steps=n_timesteps,
        max_episode_length=int(kw.pop("max_episode_length")),
        learning_rate=float(kw.pop("learning_rate")),
        nn_hidden_layer_widths=np.asarray(kw.pop("nn_hidden_layer_widths"), dtype=np.int32),
        discount_factor=float(kw.pop("gamma")),
        target_network_step=int(kw.pop("TN_step")),
        target_network_active=bool(kw.pop("target_network_active")),
        n_returns_interval=eval_interval,
        exploration_method="epsilon_greedy" if exploration_method == "egreedy" else exploration_method,
        epsilon_start=kw.pop("epsilon_start", None),
        epsilon_end=kw.pop("epsilon_end", None),
        epsilon_decay=float(kw.pop("epsilon_decay", 1.0)),
        epsilon_decay_interval=int(kw.pop("epsilon_decay_interval", 1)),
        softmax_temp=kw.pop("softmax_temp", None),
        er_active=bool(kw.pop("er_active", False)),
        er_replay_buffer_size=int(kw.pop("er_replay_buffer_size", 10000)),
        er_batch_size=int(kw.pop("er_batch_size", 64)),
        er_min_replay_size=int(kw.pop("er_min_replay_size", 100)),
        er_sample_train_frequency=int(kw.pop("er_sample_train_frequency", 1)),
        er_replay_ratio=float(kw.pop("er_replay_ratio", 1.0)),
        enable_progress_bar=True,
        emit_trial_header=False,
    )
    # Remove n_timesteps from remaining kw (already consumed)
    kw.pop("n_timesteps", None)

    rep_outputs: list[tuple[np.ndarray, np.ndarray] | None] = [None for _ in range(n_repetitions)]
    parallel_workers = max(1, min(n_repetitions, os.cpu_count() or 1))

    if parallel_workers > 1 and n_repetitions > 1:
        with ProcessPoolExecutor(max_workers=parallel_workers) as executor:
            future_to_rep = {}
            for rep in range(n_repetitions):
                run_seed = base_seed + rep
                future = executor.submit(
                    run_dqn_trial_returns,
                    seed=run_seed,
                    trial_run_index=rep + 1,
                    total_trial_runs=n_repetitions,
                    progress_bar_position=rep,
                    progress_bar_desc=f"DQN Trial {rep + 1}/{n_repetitions}",
                    **trial_common,
                )
                future_to_rep[future] = rep
            for future in as_completed(future_to_rep):
                rep = future_to_rep[future]
                rep_outputs[rep] = future.result()
    else:
        for rep in range(n_repetitions):
            run_seed = base_seed + rep
            rep_outputs[rep] = run_dqn_trial_returns(
                seed=run_seed,
                trial_run_index=rep + 1,
                total_trial_runs=n_repetitions,
                progress_bar_position=rep,
                progress_bar_desc=f"DQN Trial {rep + 1}/{n_repetitions}",
                **trial_common,
            )

    completed_rep_outputs: list[tuple[np.ndarray, np.ndarray]] = [result for result in rep_outputs if result is not None]

    returns_list = []
    timesteps = None
    for returns, rep_ts in completed_rep_outputs:
        returns_list.append(np.asarray(returns, dtype=np.float32))
        if timesteps is None:
            timesteps = np.asarray(rep_ts, dtype=np.int32)

    all_returns = np.array(returns_list)
    lc_mean = np.mean(all_returns, axis=0)
    lc_std = (np.std(all_returns, axis=0, ddof=1)
              if all_returns.shape[0] > 1 else np.zeros_like(lc_mean))
    return lc_mean, lc_std, timesteps


# ── Workbook utilities (moved from excel_workbook_utils.py) ───────────────────

SAMPLE_TEMPLATE_PATH = os.path.join("data sheets", "Sample format", "1399.03.12.xlsx")
SETTING_SHEET_PREFIX = "Setting_"

TITLE_ROW_HEIGHT = 34.25
HEADER_ROW_HEIGHT = 52.5
DATA_ROW_HEIGHT = 23.4
DEFAULT_ROW_HEIGHT = 14.5
DEFAULT_COLUMN_WIDTH = 8.90625
SHEET_ZOOM_SCALE = 85


def build_algorithm_filename(algo_name: str) -> str:
    """Return the workbook filename stem for an algorithm."""
    return str(algo_name).upper()


def _normalize_for_signature(value: Any) -> Any:
    """Normalize values so workbook values and config values can be compared reliably."""
    if isinstance(value, np.ndarray):
        return tuple(_normalize_for_signature(v) for v in value.tolist())
    if isinstance(value, (list, tuple)):
        return tuple(_normalize_for_signature(v) for v in value)
    if isinstance(value, set):
        return tuple(sorted(_normalize_for_signature(v) for v in value))
    if isinstance(value, np.bool_):
        value = bool(value)
    elif isinstance(value, (np.integer, np.floating)):
        value = value.item()
    return _value_to_text(value)


def _value_to_text(value: Any) -> str:
    """Convert a value to a compact display string suitable for Excel text cells."""
    if value is None:
        return ""
    if isinstance(value, np.ndarray):
        return str(value.tolist())
    if isinstance(value, (list, tuple)):
        return str(list(value))
    if isinstance(value, set):
        return str(sorted(value))
    if isinstance(value, (np.integer, np.floating)):
        return _fmt(value.item())
    if isinstance(value, bool):
        return "True" if value else "False"
    return str(value)


def _job_signature(job_hyperparams: dict[str, Any]) -> tuple[tuple[str, Any], ...]:
    """Create a hashable signature from a job hyperparameter dictionary."""
    return tuple(sorted((str(key), _normalize_for_signature(value)) for key, value in job_hyperparams.items()))


def _sheet_signature(sheet_hyperparams: dict[str, Any]) -> tuple[tuple[str, Any], ...]:
    """Create a hashable signature from parsed workbook hyperparameters."""
    return tuple(sorted((str(key), _normalize_for_signature(value)) for key, value in sheet_hyperparams.items()))


@lru_cache(maxsize=1)
def _load_template_sheet():
    """Load the sample formatting sheet once and reuse it for all exports."""
    if not os.path.isfile(SAMPLE_TEMPLATE_PATH):
        return None
    workbook = load_workbook(SAMPLE_TEMPLATE_PATH)
    return workbook[workbook.sheetnames[0]]


def _copy_style(source_cell, target_cell) -> None:
    """Copy all visible style attributes from one cell to another."""
    target_cell._style = copy(source_cell._style)
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.protection = copy(source_cell.protection)
        target_cell.number_format = source_cell.number_format


def _apply_excel_sheet_shell(worksheet, sheet_title: str, n_columns: int) -> None:
    """Apply sheet-level settings that match the sample workbook."""
    worksheet.sheet_view.zoomScale = SHEET_ZOOM_SCALE
    worksheet.sheet_format.defaultRowHeight = DEFAULT_ROW_HEIGHT
    worksheet.sheet_format.defaultColWidth = DEFAULT_COLUMN_WIDTH
    worksheet.freeze_panes = None
    worksheet.auto_filter.ref = None
    worksheet.row_dimensions[1].height = TITLE_ROW_HEIGHT
    worksheet.row_dimensions[2].height = HEADER_ROW_HEIGHT
    for row_index in range(3, worksheet.max_row + 1):
        worksheet.row_dimensions[row_index].height = DATA_ROW_HEIGHT


def _format_excel(
    worksheet,
    *,
    algo_name: str,
    hyperparams: dict[str, Any],
    headers: list[str],
    rows: list[list[Any]],
) -> None:
    """Format a worksheet to match the attached sample workbook style."""
    template_sheet = _load_template_sheet()
    n_columns = max(1, len(headers))

    _apply_excel_sheet_shell(worksheet, algo_name, n_columns)

    title_text = f"{str(algo_name).upper()}: " + ", ".join(
        f"{key}={_value_to_text(value)}" for key, value in hyperparams.items()
    )
    title_end = get_column_letter(n_columns)
    worksheet.merge_cells(f"A1:{title_end}1")
    worksheet["A1"] = title_text

    if template_sheet is not None:
        _copy_style(template_sheet["A1"], worksheet["A1"])
    else:
        worksheet["A1"].alignment = worksheet["A1"].alignment.copy(horizontal="center", vertical="center")
        worksheet["A1"].font = worksheet["A1"].font.copy(bold=True)

    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(2, column_index, header)
        if template_sheet is not None:
            template_col = min(column_index, template_sheet.max_column)
            _copy_style(template_sheet.cell(2, template_col), cell)
        else:
            cell.alignment = cell.alignment.copy(horizontal="center", vertical="center", wrap_text=True)
            cell.font = cell.font.copy(bold=True)

    for row_index, row_values in enumerate(rows, start=3):
        worksheet.row_dimensions[row_index].height = DATA_ROW_HEIGHT
        for column_index, value in enumerate(row_values, start=1):
            cell = worksheet.cell(row_index, column_index, value)
            if template_sheet is not None:
                template_col = min(column_index, template_sheet.max_column)
                _copy_style(template_sheet.cell(3, template_col), cell)
            else:
                cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
                cell.font = cell.font.copy(bold=True)

    if rows:
        first_data_row = rows[0]
    else:
        first_data_row = ["" for _ in headers]

    for column_index, header in enumerate(headers, start=1):
        first_value = first_data_row[column_index - 1] if column_index - 1 < len(first_data_row) else ""
        width = max(_excel_cell_display_width(header), _excel_cell_display_width(first_value)) + 1
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    worksheet.sheet_view.tabSelected = True


def _write_raw_excel_sheet(worksheet, headers: list[str], rows: list[list[Any]]) -> None:
    """Write a worksheet without merged cells or style formatting."""
    for column_index, header in enumerate(headers, start=1):
        worksheet.cell(1, column_index, header)

    for row_index, row_values in enumerate(rows, start=2):
        for column_index, value in enumerate(row_values, start=1):
            worksheet.cell(row_index, column_index, value)

    first_data_row = rows[0] if rows else []
    _set_raw_excel_column_widths(worksheet, headers, first_data_row)


def _next_setting_sheet_name(workbook) -> str:
    """Return the next available Setting_### sheet name."""
    highest_index = 0
    for sheet_name in workbook.sheetnames:
        if not sheet_name.startswith(SETTING_SHEET_PREFIX):
            continue
        suffix = sheet_name[len(SETTING_SHEET_PREFIX):]
        try:
            highest_index = max(highest_index, int(suffix))
        except ValueError:
            continue
    return f"{SETTING_SHEET_PREFIX}{highest_index + 1:03d}"


def _rows_from_result(result: tuple[np.ndarray, np.ndarray, np.ndarray]) -> list[list[Any]]:
    """Convert a result tuple into a row-wise table for Excel writing."""
    learning_curve, learning_curve_std, timesteps = result
    rows = []
    for timestep, mean_value, std_value in zip(timesteps, learning_curve, learning_curve_std):
        rows.append([timestep, mean_value, std_value])
    return rows


def _build_headers(job_hyperparams: dict[str, Any]) -> list[str]:
    """Build the Excel column headers for a saved setting."""
    return [
        "timestep",
        "learning_curve_mean",
        "learning_curve_std",
        *[str(key) for key in job_hyperparams.keys()],
        "curve_label",
    ]


def _build_rows(job_hyperparams: dict[str, Any], curve_label: str, result: tuple[np.ndarray, np.ndarray, np.ndarray]) -> list[list[Any]]:
    """Build the Excel rows for a single setting sheet."""
    base_rows = _rows_from_result(result)
    hyperparam_values = [_value_to_text(value) for value in job_hyperparams.values()]
    rows: list[list[Any]] = []
    for timestep, mean_value, std_value in base_rows:
        rows.append([timestep, mean_value, std_value, *hyperparam_values, curve_label])
    return rows


def _parse_sheet_entry(worksheet, *, formatted_sheets: bool = False) -> dict[str, Any] | None:
    """Parse one worksheet into the in-memory result format."""
    header_row = 2 if formatted_sheets else 1
    data_start_row = 3 if formatted_sheets else 2

    headers: list[str] = []
    header_to_column: dict[str, int] = {}

    for column_index in range(1, worksheet.max_column + 1):
        header_value = worksheet.cell(header_row, column_index).value
        if header_value is None:
            continue
        header_text = str(header_value)
        if header_text in header_to_column:
            continue
        header_to_column[header_text] = column_index
        headers.append(header_text)

    required = {"timestep", "learning_curve_mean", "learning_curve_std"}
    if not required.issubset(set(headers)):
        return None

    column_data: dict[str, list[Any]] = {header: [] for header in headers}
    for row_index in range(data_start_row, worksheet.max_row + 1):
        non_empty = False
        for header in headers:
            column_index = header_to_column[header]
            value = worksheet.cell(row_index, column_index).value
            if value is not None:
                non_empty = True
            column_data[header].append(value)
        if not non_empty:
            continue

    try:
        timesteps = np.asarray(column_data["timestep"], dtype=np.int32)
        learning_curve = np.asarray(column_data["learning_curve_mean"], dtype=np.float32)
        learning_curve_std = np.asarray(column_data["learning_curve_std"], dtype=np.float32)
    except Exception:
        return None

    hyperparams: dict[str, Any] = {}
    for header in headers:
        if header in required or header == "curve_label":
            continue
        values = column_data[header]
        first_value = values[0] if values else None
        if first_value is not None:
            hyperparams[header] = _value_to_text(first_value)

    curve_values = column_data.get("curve_label", [])
    curve_label = next((str(value) for value in curve_values if value not in (None, "")), "")

    return {
        "learning_curve": learning_curve,
        "learning_curve_std": learning_curve_std,
        "timesteps": timesteps,
        "curve_label": curve_label,
        "hyperparams": hyperparams,
    }


def load_algorithm_workbook(
    filepath: str,
    *,
    setting_jobs: list[dict[str, Any]] | None = None,
    formatted_sheets: bool = False,
) -> tuple[list[dict[str, Any] | None], dict[str, tuple[Any, Any]]]:
    """Load matching sheets from a workbook."""
    try:
        workbook = load_workbook(filepath)
    except Exception as exc:
        raise ValueError(f"Failed to read Excel file '{filepath}': {exc}") from exc

    parsed_entries: list[dict[str, Any] | None] = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        entry = _parse_sheet_entry(worksheet, formatted_sheets=formatted_sheets)
        if entry is None:
            entry = _parse_sheet_entry(worksheet, formatted_sheets=not formatted_sheets)
        if entry is not None:
            parsed_entries.append(entry)

    if setting_jobs is None:
        return parsed_entries, {}

    job_map: dict[tuple[tuple[str, Any], ...], int] = {}
    for index, job in enumerate(setting_jobs):
        job_map[_job_signature(job["hyperparams"])] = index

    aligned_entries: list[dict[str, Any] | None] = [None] * len(setting_jobs)
    for entry in parsed_entries:
        signature = _sheet_signature(entry["hyperparams"])
        job_index = job_map.get(signature)
        if job_index is not None and aligned_entries[job_index] is None:
            aligned_entries[job_index] = entry

    return aligned_entries, {}


def save_algorithm_workbook(
    dir_path: str,
    base_filename: str,
    algo_name: str,
    setting_jobs: list[dict[str, Any]],
    setting_results: list[tuple[np.ndarray, np.ndarray, np.ndarray] | None],
    format_sheets: bool = False,
) -> str:
    """Save or append settings into an algorithm workbook."""
    os.makedirs(dir_path, exist_ok=True)
    filepath = os.path.join(dir_path, f"{base_filename}.xlsx")

    existing_entries: list[dict[str, Any]] = []
    if os.path.isfile(filepath):
        try:
            existing_entries, _ = load_algorithm_workbook(filepath, formatted_sheets=format_sheets)
        except Exception:
            existing_entries = []

    existing_signatures: set[tuple[tuple[str, Any], ...]] = set()
    for entry in existing_entries:
        existing_signatures.add(_sheet_signature(entry["hyperparams"]))

    all_entries = list(existing_entries)
    added_count = 0
    for job, result in zip(setting_jobs, setting_results):
        if result is None:
            continue

        job_signature = _job_signature(job["hyperparams"])
        if job_signature in existing_signatures:
            continue

        learning_curve, learning_curve_std, timesteps = result
        all_entries.append({
            "learning_curve": learning_curve,
            "learning_curve_std": learning_curve_std,
            "timesteps": timesteps,
            "curve_label": job["curve_label"],
            "hyperparams": job["hyperparams"],
        })
        existing_signatures.add(job_signature)
        added_count += 1

    workbook = Workbook()
    default_sheet = workbook.active
    if default_sheet is not None:
        workbook.remove(default_sheet)

    if not all_entries:
        worksheet = workbook.create_sheet(title=_next_setting_sheet_name(workbook))
        if format_sheets:
            _format_excel(
                worksheet,
                algo_name=algo_name,
                hyperparams={},
                headers=["timestep", "learning_curve_mean", "learning_curve_std", "curve_label"],
                rows=[],
            )
        else:
            _write_raw_excel_sheet(
                worksheet,
                ["timestep", "learning_curve_mean", "learning_curve_std", "curve_label"],
                [],
            )
    else:
        for index, entry in enumerate(all_entries, start=1):
            worksheet = workbook.create_sheet(title=f"Setting_{index:03d}")
            headers = _build_headers(entry["hyperparams"])
            rows = _build_rows(entry["hyperparams"], entry["curve_label"], (
                entry["learning_curve"],
                entry["learning_curve_std"],
                entry["timesteps"],
            ))
            if format_sheets:
                _format_excel(
                    worksheet,
                    algo_name=algo_name,
                    hyperparams=entry["hyperparams"],
                    headers=headers,
                    rows=rows,
                )
            else:
                _write_raw_excel_sheet(worksheet, headers, rows)

    workbook.save(filepath)
    print(f"Saved {added_count} new setting(s) to {filepath}")
    return filepath
