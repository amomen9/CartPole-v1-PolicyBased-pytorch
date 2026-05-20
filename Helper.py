#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Practical for course 'Reinforcement Learning',
Leiden University, The Netherlands
By Thomas Moerland
"""

"""
Facilitation functions for the experiment pipeline.
Contains file/Excel utility helpers and algorithm job builders.
"""

import os
import shutil
import time
import ast
import glob
from copy import copy
from datetime import datetime
from functools import lru_cache
from concurrent.futures import ProcessPoolExecutor, as_completed
from typing import Any, Tuple
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
from tqdm import tqdm
from multiprocessing import Manager
import matplotlib.pyplot as plt
import matplotlib.pyplot as visplt
from scipy.signal import savgol_filter
from scipy.stats import t as t_dist


# Timestamp captured at module import time - represents the start of execution
# for the current run. Used to suffix plot filenames so all plots from one run
# share the same timestamp (format: YYYY.MM.DD_HH.MM.SS).
RUN_TIMESTAMP = datetime.now().strftime("%Y.%m.%d_%H.%M.%S")


def _create_step_progress_bar(total, desc, position=None, leave=True):
    """Create a tqdm progress bar with the shared project formatting."""

    tqdm_kwargs = {
        "total": total,
        "desc": desc,
        "unit": "step",
        "bar_format": "{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]",
        "dynamic_ncols": True,
        "leave": leave,
    }
    if position is not None:
        tqdm_kwargs["position"] = int(position)
    return tqdm(**tqdm_kwargs)

# Begin Class LearningCurvePlot ##############################################################
class LearningCurvePlot:

    def __init__(self,title=None):
        self.fig,self.ax = plt.subplots()
        self.ax.set_xlabel('Timestep')
        self.ax.set_ylabel('Episode Return')
        if title is not None:
            self.ax.set_title(title)

    def add_curve(self, x, y, label=None, ls="solid", color=None):
        ''' y: vector of average reward results
        label: string to appear as label in plot legend '''
        plot_kwargs = {"ls": ls}
        if color is not None:
            plot_kwargs["color"] = color
        if label is not None:
            self.ax.plot(x, y, label=label, **plot_kwargs)
        else:
            self.ax.plot(x, y, **plot_kwargs)

    def add_shaded_ci(self, x, y_mean, y_std, n, alpha=0.2, fill_opacity=0.15, y_lower_cap=None, y_upper_cap=None, color=None):
        '''Add a shaded confidence band around the mean curve.
        alpha controls CI significance (e.g., 0.05 for 95% CI),
        fill_opacity controls the visual transparency of the shaded area.'''
        t_crit = t_dist.ppf(1 - alpha / 2, df=max(n - 1, 1))
        margin = t_crit * y_std / np.sqrt(max(n, 1))
        y_lower = y_mean - margin
        y_upper = y_mean + margin
        if y_lower_cap is not None:
            y_lower = np.maximum(y_lower, y_lower_cap)
        if y_upper_cap is not None:
            y_upper = np.minimum(y_upper, y_upper_cap)
        if color is None:
            color = self.ax.get_lines()[-1].get_color()  # match the last plotted line
        self.ax.fill_between(x, y_lower, y_upper,
                             alpha=fill_opacity, color=color)

    def set_ylim(self,lower,upper):
        self.ax.set_ylim([lower,upper])

    def add_hline(self,height,label):
        self.ax.axhline(height,ls='--',c='k',label=label)

    @staticmethod
    def _resolve_non_overwriting_path(output_path: str) -> str:
        """Return output_path if it doesn't exist; otherwise add (1), (2), ... before the extension."""
        if not os.path.exists(output_path):
            return output_path

        root, ext = os.path.splitext(output_path)
        i = 1
        while True:
            candidate = f"{root} ({i}){ext}"
            if not os.path.exists(candidate):
                return candidate
            i += 1

    def save(self,name='test.png') -> str:
        ''' name: string for filename of saved figure.

        If the target filename already exists, saves to an enumerated name:
        file.png -> file (1).png -> file (2).png -> ...
        '''
        self.ax.legend(
            fontsize=8,
            handlelength=1.2,
            handletextpad=0.4,
            borderpad=0.25,
            labelspacing=0.25,
            borderaxespad=0.3,
        )
        self.fig.tight_layout()

        output_path = name
        if not os.path.isabs(name):
            os.makedirs("plots", exist_ok=True)
            output_path = os.path.join("plots", os.path.basename(name))
        else:
            # Ensure parent folder exists for absolute paths
            out_dir = os.path.dirname(output_path)
            if out_dir:
                os.makedirs(out_dir, exist_ok=True)

        root, ext = os.path.splitext(output_path)
        output_path = f"{root}_{RUN_TIMESTAMP}{ext}"

        output_path = self._resolve_non_overwriting_path(output_path)
        self.fig.savefig(output_path,dpi=300)
        return output_path
# End Class LearningCurvePlot ##############################################################



def smooth(y, window, poly=2):
    '''
    y: vector to be smoothed
    window: size of the smoothing window '''
    return savgol_filter(y,window,poly)


def _apply_optional_smoothing(learning_curve, plot_smoothing_window):
    if plot_smoothing_window is None:
        return learning_curve
    max_window = len(learning_curve) if len(learning_curve) % 2 == 1 else len(learning_curve) - 1
    window = min(int(plot_smoothing_window), max_window)
    if window >= 3:
        return smooth(learning_curve, window)
    return learning_curve


def _load_benchmark_curve(
    benchmark_curve,
    project_eval_interval,
    project_n_timesteps,
    benchmark_eval_interval=250,
    episode_return_column="Episode_Return",
    benchmark_name="Baseline",
):
    benchmark_files = {
        1: os.path.join("Baseline data", "BaselineDataCartPole_run1.csv"),
        2: os.path.join("Baseline data", "BaselineDataCartPole_run2.csv"),
    }
    if benchmark_curve not in benchmark_files:
        raise ValueError("benchmark_curve must be 1 or 2.")

    benchmark_path = benchmark_files[benchmark_curve]
    if not os.path.isfile(benchmark_path):
        LAST_BENCHMARK_WAS_MISSING = True
        print(
            f"!!!Warning! {benchmark_name} files were not found on the disk, continuing the operation without {benchmark_name} (The plots will not include the {benchmark_name}.)!!!\n\n"
        )
        return np.array([], dtype=np.int32), np.array([], dtype=np.float32)

    LAST_BENCHMARK_WAS_MISSING = False

    data = np.genfromtxt(benchmark_path, delimiter=",", names=True)
    if data.dtype.names is None or episode_return_column not in data.dtype.names:
        raise ValueError(
            f"Selected benchmark file does not contain requested column '{episode_return_column}'."
        )

    env_steps = np.atleast_1d(np.asarray(data["env_step"], dtype=np.float32))
    returns = np.atleast_1d(np.asarray(data[episode_return_column], dtype=np.float32))

    if env_steps.size == 0 or returns.size == 0:
        raise ValueError("Selected benchmark file has no usable data.")

    valid_rows = np.isfinite(env_steps) & np.isfinite(returns)
    env_steps = env_steps[valid_rows]
    returns = returns[valid_rows]

    if env_steps.size == 0:
        raise ValueError("Selected benchmark file contains only invalid rows.")

    sort_idx = np.argsort(env_steps)
    env_steps = env_steps[sort_idx]
    returns = returns[sort_idx]

    if project_eval_interval != benchmark_eval_interval and env_steps.size >= 2:
        normalized_steps = np.arange(
            env_steps[0],
            env_steps[-1] + project_eval_interval,
            project_eval_interval,
            dtype=np.float32,
        )
        normalized_steps = normalized_steps[normalized_steps <= env_steps[-1]]
        returns = np.interp(normalized_steps, env_steps, returns).astype(np.float32)
        env_steps = normalized_steps

    in_horizon = env_steps <= float(project_n_timesteps)
    if not np.any(in_horizon):
        print(
            f"[benchmark] No benchmark points fall within project_n_timesteps={project_n_timesteps}; "
            "using the full benchmark curve instead."
        )
        return env_steps.astype(np.int32), returns

    env_steps = env_steps[in_horizon]
    returns = returns[in_horizon]
    return env_steps.astype(np.int32), returns


def average_over_repetitions(
    method,
    n_repetitions,
    n_timesteps,
    eval_interval,
    max_episode_length,
    actor_lr,
    gamma,
    actor_hidden_nn=np.array([16, 16]),
    critic_hidden_nn=np.array([64, 64]),
    critic_lr=0.001,
    base_seed=42,
    plot_smoothing_window=None,
    eval_with_env_episode_trials: bool = False,
    n_eval_episodes: int = 5,
    return_raw=False,
    unused_cpu_cores: int = 0,
):
    """Run 'n_repetitions' of the given method and return (mean, std, timesteps)."""

    from concurrent.futures import ProcessPoolExecutor, as_completed  # noqa: F401
    from multiprocessing import Manager
    import time as _time

    # Lazy import to avoid circular import on module load.
    from Library import _run_single_repetition

    returns_over_repetitions = []
    timesteps = None

    cpu_count = os.cpu_count() or 1
    if unused_cpu_cores is None:
        unused_cpu_cores = 0
    unused_cpu_cores = int(unused_cpu_cores)
    if unused_cpu_cores < 0:
        unused_cpu_cores = 0

    available_cpus = max(1, cpu_count - unused_cpu_cores)
    parallel_workers = max(1, min(n_repetitions, available_cpus))
    use_parallel = parallel_workers > 1 and n_repetitions > 1

    if use_parallel:
        manager = Manager()
        step_counters = [manager.Value("i", 0) for _ in range(n_repetitions)]
        try:
            with ProcessPoolExecutor(max_workers=parallel_workers) as executor:
                future_to_rep = {}
                for rep in range(n_repetitions):
                    run_seed = base_seed + rep
                    future = executor.submit(
                        _run_single_repetition,
                        method=method,
                        actor_hidden_nn=actor_hidden_nn,
                        critic_hidden_nn=critic_hidden_nn,
                        actor_lr=actor_lr,
                        critic_lr=critic_lr,
                        gamma=gamma,
                        max_episode_length=max_episode_length,
                        n_timesteps=n_timesteps,
                        eval_interval=eval_interval,
                        run_seed=run_seed,
                        rep_index=rep,
                        n_repetitions=n_repetitions,
                        enable_progress_bar=False,
                        shared_step_counter=step_counters[rep],
                        eval_with_env_episode_trials=eval_with_env_episode_trials,
                        n_eval_episodes=n_eval_episodes,
                    )
                    future_to_rep[future] = rep

                max_visible_bars = min(
                    n_repetitions,
                    int(os.environ.get("RL_MAX_TQDM_BARS", "10")),
                )
                n_groups = max(1, max_visible_bars)

                group_size = (n_repetitions + n_groups - 1) // n_groups

                groups: list[list[int]] = []
                for gi in range(n_groups):
                    start = gi * group_size
                    end = min((gi + 1) * group_size, n_repetitions)
                    if start >= end:
                        break
                    groups.append(list(range(start, end)))

                rep_to_group: dict[int, int] = {}
                for gi, group in enumerate(groups):
                    for rep in group:
                        rep_to_group[rep] = gi

                pbars = {}
                for gi, group in enumerate(groups):
                    rep_list_str = ",".join(str(rep + 1) for rep in group)
                    pbars[gi] = _create_step_progress_bar(
                        total=int(n_timesteps) * len(group),
                        desc=f"{method.upper()} Rep {rep_list_str}/{len(groups)}",
                        position=gi,
                        leave=True,
                    )

                rep_last = [0 for _ in range(n_repetitions)]

                done_futures = set()
                try:
                    while len(done_futures) < n_repetitions:
                        # Update merged/grouped bars as the SUM of rep progress.
                        # Since each group's total is n_timesteps * (num_reps_in_group),
                        # progress fills more slowly when multiple reps are merged.
                        for gi, group in enumerate(groups):
                            sum_delta = 0
                            for rep in group:
                                current = step_counters[rep].value
                                delta = current - rep_last[rep]
                                if delta > 0:
                                    sum_delta += delta
                                    rep_last[rep] = current
                            if sum_delta > 0:
                                pbars[gi].update(sum_delta)

                        for future in list(future_to_rep):
                            if future not in done_futures and future.done():
                                done_futures.add(future)
                                rep = future_to_rep[future]
                                rep_returns, rep_timesteps = future.result()
                                returns_over_repetitions.append(
                                    np.asarray(rep_returns, dtype=np.float32)
                                )
                                if timesteps is None:
                                    timesteps = np.asarray(rep_timesteps, dtype=np.int32)

                                gid = rep_to_group.get(rep)
                                if gid is not None:
                                    current = step_counters[rep].value
                                    delta = current - rep_last[rep]
                                    if delta > 0:
                                        pbars[gid].update(delta)
                                        rep_last[rep] = current

                        _time.sleep(0.25)
                finally:
                    for pb in pbars.values():
                        pb.close()
                    print()
        finally:
            manager.shutdown()
    else:
        for rep in range(n_repetitions):
            run_seed = base_seed + rep
            rep_returns, rep_timesteps = _run_single_repetition(
                method=method,
                actor_hidden_nn=actor_hidden_nn,
                critic_hidden_nn=critic_hidden_nn,
                actor_lr=actor_lr,
                critic_lr=critic_lr,
                gamma=gamma,
                max_episode_length=max_episode_length,
                n_timesteps=n_timesteps,
                eval_interval=eval_interval,
                run_seed=run_seed,
                rep_index=rep,
                n_repetitions=n_repetitions,
                enable_progress_bar=True,
                eval_with_env_episode_trials=eval_with_env_episode_trials,
                n_eval_episodes=n_eval_episodes,
            )
            returns_over_repetitions.append(np.asarray(rep_returns, dtype=np.float32))
            if timesteps is None:
                timesteps = np.asarray(rep_timesteps, dtype=np.int32)

    min_length = min(len(rep_returns) for rep_returns in returns_over_repetitions)
    returns_over_repetitions = [
        np.asarray(rep_returns[:min_length], dtype=np.float32)
        for rep_returns in returns_over_repetitions
    ]
    if timesteps is not None:
        timesteps = np.asarray(timesteps[:min_length], dtype=np.int32)

    all_returns = np.array(returns_over_repetitions)
    learning_curve = np.mean(all_returns, axis=0)
    learning_curve_std = (
        np.std(all_returns, axis=0, ddof=1)
        if all_returns.shape[0] > 1
        else np.zeros_like(learning_curve)
    )
    learning_curve = _apply_optional_smoothing(learning_curve, plot_smoothing_window)
    learning_curve_std = _apply_optional_smoothing(learning_curve_std, plot_smoothing_window)

    if return_raw:
        raw_returns = np.asarray(returns_over_repetitions, dtype=np.float32)
        return learning_curve, learning_curve_std, timesteps, raw_returns
    return learning_curve, learning_curve_std, timesteps


### One suggested simplest policy {eps, 1-eps} is below, however, I implement the one that was mentioned in the assignment instead.
#def egreedy(Qa_s, eps):
#    ''' Qa_s: vector of action values for state s
#        epsilon: exploration parameter '''
#    if np.random.rand() < eps:
#        return np.random.randint(0,len(Qa_s)) # Explore action space
#    else:
#        return argmax(Qa_s) # Exploit learned values

def egreedy(Qa_s, eps):
    """
    Sample one action using epsilon-greedy policy
    Qa_s: 1D array of Q-values for current state's actions
    eps: epsilon in the closed boundary [0,1]
    """
    n_A = len(Qa_s)     # number of actions
    greedy_a = argmax(Qa_s)  # tie breaking argmax()
    # Base probability for all actions, fill probs matrix with the same values (will not sum up to 1 yet)
    probs = np.full(n_A, eps / n_A, dtype=float)
    # Greedy action gets the remaining probability mass (1 - eps) plus its share of the exploration probability (eps/n_A)
    probs[greedy_a] = 1.0 - eps * (n_A - 1) / n_A
    selected_action = np.random.choice(n_A, p=probs)
    # Sample action from this distribution
    return selected_action


def argmax(x):
    ''' Own variant of np.argmax with random tie breaking '''
    try:
        return np.random.choice(np.where(x == np.max(x))[0])
    except:
        return np.argmax(x)

def linear_anneal(t,T,start,final,percentage):
    ''' Linear annealing scheduler
    t: current timestep
    T: total timesteps
    start: initial value
    final: value after percentage*T steps
    percentage: percentage of T after which annealing finishes
    '''
    final_from_T = int(percentage*T)
    if t > final_from_T:
        return final
    else:
        return final + (start - final) * (final_from_T - t)/final_from_T

################[ Policy_NN Class              ]################
def softmax(x, temp):   # aka Boltzmann policy (Mentioned in Assignment 1 as Boltzmann in the assignment)
    ''' Computes the softmax of vector x with temperature parameter 'temp' '''
    x = x / temp # scale by temperature
    z = x - max(x) # substract max to prevent overflow of softmax
    probs = np.exp(z)/np.sum(np.exp(z)) # compute softmax
    selected_action = np.random.choice(len(x), p=probs) # Sample action from
    return int(selected_action)
####################################################################



if __name__ == '__main__':
    # Test Learning curve plot
    # x = np.arange(100)
    # y = 0.01*x + np.random.rand(100) - 0.4 # generate some learning curve y
    # LCTest = LearningCurvePlot(title="Test Learning Curve")
    # LCTest.add_curve(x,y,label='method 1')
    # LCTest.add_curve(x,smooth(y,window=35),label='method 1 smoothed')
    # LCTest.save(name='learning_curve_test.png')
    # import Environment
    # env = Environment.CartPoleEnvironment(max_episode_length=500, render_mode="rgb_array")
    # CartPole_plot = CartPoleAgentPlot(env, title="Test CartPole Agent Plot", plot=False)
    # anim = CartPole_plot.test_episode(env, policy="test_policy")
    # visplt.show()
    pass





# ── File / Excel utilities ────────────────────────────────────────────────────

def _fmt(v):
    """Format numbers safely for filenames."""
    if isinstance(v, float):
        return f"{v:.3g}".replace(".", "p")
    return str(v)


def _format_legend_value(val: Any) -> str:
    """Format legend values, unwrapping single-item containers."""
    if isinstance(val, np.ndarray):
        values = np.atleast_1d(val).tolist()
    elif isinstance(val, (list, tuple, set)):
        values = list(val)
    elif hasattr(val, "__iter__") and not isinstance(val, (str, bytes, dict)):
        values = list(val)
    else:
        return _fmt_legend(val)

    if len(values) == 0:
        return "[]"
    if len(values) == 1:
        return _fmt_legend(values[0])
    return "[" + ",".join(_fmt_legend(item) for item in values) + "]"


def _format_legend_label(label: str) -> str:
    """Return legend labels exactly as configured in Experiment.py."""
    return str(label)


def _excel_cell_display_width(value: Any) -> int:
    """Return the visible width of a value when exported to Excel."""
    if value is None:
        return 0
    if isinstance(value, (float, np.floating)) and np.isnan(value):
        return 0
    text = str(value)
    if not text:
        return 0
    return max(len(line) for line in text.splitlines())


def _autosize_excel_columns(worksheet, dataframe) -> None:
    """Resize Excel columns so their contents fit the widest cell in each column."""
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        max_width = _excel_cell_display_width(column_name)
        for value in dataframe[column_name].tolist():
            max_width = max(max_width, _excel_cell_display_width(value))
        worksheet.column_dimensions[get_column_letter(column_index)].width = max_width


def _set_raw_excel_column_widths(worksheet, headers, first_row_values) -> None:
    """Set raw-sheet widths from the header and first data row only."""
    if first_row_values is None:
        first_row_values = []
    for column_index, header in enumerate(headers, start=1):
        first_value = first_row_values[column_index - 1] if column_index - 1 < len(first_row_values) else ""
        width = max(_excel_cell_display_width(header), _excel_cell_display_width(first_value)) + 1
        worksheet.column_dimensions[get_column_letter(column_index)].width = width


def _sheet_hp_text(value: Any) -> str:
    """Return a stable string representation for Excel hyperparameter values."""
    if value is None:
        return ""
    if isinstance(value, np.ndarray):
        if value.ndim == 0:
            value = value.item()
        else:
            value = value.tolist()
    if isinstance(value, (list, tuple, set)):
        items = list(value)
        if len(items) == 0:
            return "[]"
        if len(items) == 1:
            return _sheet_hp_text(items[0])
        return "[" + ",".join(_sheet_hp_text(item) for item in items) + "]"
    if isinstance(value, np.bool_):
        value = bool(value)
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (np.integer, int)):
        return str(int(value))
    if isinstance(value, (np.floating, float)):
        number = float(value)
        if number.is_integer():
            return str(int(number))
        return format(number, ".6g")
    if isinstance(value, str):
        text = value.strip()
        if text.startswith("[") and text.endswith("]"):
            try:
                parsed = ast.literal_eval(text)
                return _sheet_hp_text(parsed)
            except Exception:
                return text.replace(", ", ",")
        return text
    return str(value).strip()


def _sheet_hp_text_candidates(value: Any) -> list[str]:
    """Return the normalized text candidate(s) for a config value."""
    return [_sheet_hp_text(value)]


def _empty_data_sheets_dir(dir_path):
    if os.path.isdir(dir_path):
        shutil.rmtree(dir_path)
    os.makedirs(dir_path, exist_ok=True)


def _show_excel_save_permission_warning(filepath: str) -> None:
    """Show a blocking warning figure when an Excel file is locked or write-protected."""
    warning_fig = plt.figure(figsize=(12, 6))
    warning_fig.patch.set_facecolor("#fff3cd")
    ax = warning_fig.add_subplot(111)
    ax.axis("off")
    ax.text(
        0.5,
        0.68,
        "Permission denied",
        ha="center",
        va="center",
        fontsize=28,
        fontweight="bold",
        color="#b00020",
        transform=ax.transAxes,
    )
    ax.text(
        0.5,
        0.42,
        f"Could not save:\n{filepath}\n\nClose the file in Excel or fix the permissions,\nthen close this window to retry saving.",
        ha="center",
        va="center",
        fontsize=18,
        color="black",
        transform=ax.transAxes,
    )
    warning_fig.tight_layout()
    plt.show(block=True)
    plt.close(warning_fig)


def _save_results_to_excel(dir_path, base_filename, setting_jobs, setting_results):
    filepath = os.path.join(dir_path, f"{base_filename}.xlsx")
    try:
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            for idx, (job, result) in enumerate(zip(setting_jobs, setting_results)):
                learning_curve, learning_curve_std, timesteps = result
                sheet_name = f"Setting_{idx + 1:03d}"
                data = {
                    "timestep": timesteps,
                    "learning_curve_mean": learning_curve,
                    "learning_curve_std": learning_curve_std,
                }
                hp = job["hyperparams"]
                for key, value in hp.items():
                    data[key] = value
                data["curve_label"] = job["curve_label"]
                df = pd.DataFrame(data)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                _autosize_excel_columns(writer.sheets[sheet_name], df)
    except PermissionError:
        _show_excel_save_permission_warning(filepath)
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            for idx, (job, result) in enumerate(zip(setting_jobs, setting_results)):
                learning_curve, learning_curve_std, timesteps = result
                sheet_name = f"Setting_{idx + 1:03d}"
                data = {
                    "timestep": timesteps,
                    "learning_curve_mean": learning_curve,
                    "learning_curve_std": learning_curve_std,
                }
                hp = job["hyperparams"]
                for key, value in hp.items():
                    data[key] = value
                data["curve_label"] = job["curve_label"]
                df = pd.DataFrame(data)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                _autosize_excel_columns(writer.sheets[sheet_name], df)
    print(f"Saved {len(setting_results)} settings to {filepath}")


def _load_results_from_excel(
    filepath,
    algo_config: dict[str, Any] | None,
    formatted_sheets: bool = False,
    *,
    global_config: dict[str, Any] | None = None,
    original_algo_config: dict[str, Any] | None = None,
):
    """Load results from an Excel file, validating each sheet's hyperparameters
    against the current algo config and global_config.

    Workbook layout assumption:
    - Row 1 is the title / merged heading row.
    - Row 2 contains the actual column headers.
    - Data starts on row 3.

    Validation is per-sheet: each sheet must agree with every key in the filtered
    algo_config and filtered global_config (a missing column on the sheet is
    treated as a match - lenient back-compat for workbooks written before
    global_config was persisted per sheet). The 'n_timesteps' key is the
    exception, validated with '>=' rather than equality.

    Returns a tuple (results, mismatches):
        results: list of dicts [{"learning_curve", "learning_curve_std", "timesteps", "curve_label"}, ...]
        mismatches: dict {param_name: (sheet_value, config_value)} for the first mismatch
                    encountered per parameter across all skipped sheets.
    Sheets that don't match are skipped.
    """

    del original_algo_config  # accepted for caller compatibility; no workbook-level check.

    algo_config = algo_config or {}
    global_filtered = _meta_filtered_items(global_config, GLOBAL_CONFIG_EXCLUSIONS) if global_config else {}

    running_n_timesteps: int | None = None
    if global_config is not None:
        try:
            running_n_timesteps = int(float(global_config.get(GLOBAL_CONFIG_NTIMESTEPS_KEY)))
        except (TypeError, ValueError):
            running_n_timesteps = None

    header_row = 1 if formatted_sheets else 0
    try:
        sheets = pd.read_excel(filepath, sheet_name=None, engine="openpyxl", header=header_row)
    except Exception as exc:
        raise ValueError(f"Failed to read Excel file '{filepath}': {exc}") from exc

    def _extract_matching_results(sheet_map):
        basename = os.path.basename(filepath)
        algo_prefix = os.path.splitext(basename)[0].upper()
        legend: dict[str, tuple[str, bool]] = _resolve_legend_flags(algo_config or {}, warn_on_suppression=False)
        skip_keys = ALGO_CONFIG_EXCLUSIONS

        extracted_results = []
        extracted_mismatches = {}
        required_columns = {"timestep", "learning_curve_mean", "learning_curve_std"}

        for sheet_name in sorted(sheet_map.keys()):
            df = sheet_map[sheet_name]
            if not required_columns.issubset(set(df.columns)):
                continue

            sheet_matched = True
            if algo_config:
                for cfg_key, cfg_val in algo_config.items():
                    if cfg_key in skip_keys:
                        continue
                    if cfg_key not in df.columns:
                        # Only invalidate reuse when the current run depends on these eval-mode keys.
                        if cfg_key == "eval_with_env_episode_trials":
                            if bool(cfg_val) is True:
                                sheet_matched = False
                                if cfg_key not in extracted_mismatches:
                                    extracted_mismatches[cfg_key] = ("<missing_column>", cfg_val)
                        elif cfg_key == "n_eval_episodes":
                            if bool(algo_config.get("eval_with_env_episode_trials", False)) is True:
                                sheet_matched = False
                                if cfg_key not in extracted_mismatches:
                                    extracted_mismatches[cfg_key] = ("<missing_column>", cfg_val)
                        continue

                    sheet_val = df[cfg_key].iloc[0]
                    sheet_val_text = _sheet_hp_text(_parse_sheet_value(sheet_val))
                    cfg_val_candidates = _sheet_hp_text_candidates(cfg_val)

                    if sheet_val_text not in cfg_val_candidates:
                        sheet_matched = False
                        if cfg_key not in extracted_mismatches:
                            extracted_mismatches[cfg_key] = (sheet_val_text, cfg_val)

            if sheet_matched and global_filtered:
                for gc_key, gc_text in global_filtered.items():
                    if gc_key not in df.columns:
                        continue  # lenient back-compat: pre-existing sheets without global_config columns are accepted
                    sheet_val_text = _sheet_hp_text(_parse_sheet_value(df[gc_key].iloc[0]))
                    if gc_key == GLOBAL_CONFIG_NTIMESTEPS_KEY:
                        try:
                            if int(float(sheet_val_text)) >= int(float(gc_text)):
                                continue
                        except (TypeError, ValueError):
                            pass
                        sheet_matched = False
                        tagged = f"global:{gc_key}"
                        if tagged not in extracted_mismatches:
                            extracted_mismatches[tagged] = (sheet_val_text, f"{gc_text} (need disk >= running)")
                        continue
                    if sheet_val_text != gc_text:
                        sheet_matched = False
                        tagged = f"global:{gc_key}"
                        if tagged not in extracted_mismatches:
                            extracted_mismatches[tagged] = (sheet_val_text, gc_text)

            if not sheet_matched:
                continue

            try:
                timesteps = df["timestep"].values.astype(np.int32)
                learning_curve = df["learning_curve_mean"].values.astype(np.float32)
                learning_curve_std = df["learning_curve_std"].values.astype(np.float32)
            except Exception:
                continue

            if running_n_timesteps is not None:
                mask = timesteps <= running_n_timesteps
                timesteps = timesteps[mask]
                learning_curve = learning_curve[mask]
                learning_curve_std = learning_curve_std[mask]

            label_parts = [algo_prefix]
            for legend_key, (legend_label, show) in legend.items():
                if not show:
                    continue
                if legend_key in df.columns:
                    val = _parse_sheet_value(df[legend_key].iloc[0])
                elif legend_key in algo_config:
                    val = algo_config[legend_key]
                else:
                    continue

                legend_label = _format_legend_label(legend_label)

                if isinstance(val, bool):
                    label_parts.append(f"{legend_label}{val}")
                else:
                    label_parts.append(f"{legend_label}{_format_legend_value(val)}")

            extracted_results.append({
                "learning_curve": learning_curve,
                "learning_curve_std": learning_curve_std,
                "timesteps": timesteps,
                "curve_label": ", ".join(label_parts),
            })

        return extracted_results, extracted_mismatches

    results, mismatches = _extract_matching_results(sheets)

    # Fallback for any older workbook layout that uses the opposite header row.
    if not results:
        try:
            fallback_header = 0 if formatted_sheets else 1
            fallback_sheets = pd.read_excel(filepath, sheet_name=None, engine="openpyxl", header=fallback_header)
        except Exception:
            return results, mismatches
        fallback_results, fallback_mismatches = _extract_matching_results(fallback_sheets)
        if fallback_results:
            return fallback_results, fallback_mismatches

    return results, mismatches


def _parse_sheet_value(val):
    """Parse a value read from an Excel HP column into a stable string."""
    return _sheet_hp_text(val)


def _match_sheets_to_jobs(
    filepath: str,
    setting_jobs: list[dict[str, Any]],
    *,
    formatted_sheets: bool = False,
    global_config: dict[str, Any] | None = None,
    algo_config: dict[str, Any] | None = None,
) -> tuple[list[dict[str, Any] | None], dict[str, tuple[Any, Any]]]:
    """For every job, scan every sheet in the workbook for the first match.

    A sheet matches a job when every non-excluded key in the job's combined
    hyperparams (its own 'hyperparams' plus the relevant 'global_config'
    keys) agrees with the corresponding sheet column. 'n_timesteps' is
    matched with '>=' (the learning curve is cropped to the running value).
    Lenient back-compat: if the sheet is missing 'eval_with_env_episode_trials'
    the job still matches when its value is False; similarly for
    'n_eval_episodes' when env-episode evaluation is disabled. Sheet columns
    not referenced by the job are ignored. Each sheet matches at most one job
    (first-job-first-match wins).

    Returns (aligned, mismatches) where 'aligned[i]' is 'None' when job
    'i' has no match, otherwise a result dict with the cropped learning
    curve, std, timesteps and a curve label rebuilt from the current legend.
    'mismatches' aggregates the first informative reason per parameter that
    blocked any candidate match.
    """
    algo_config = algo_config or {}
    global_filtered = (
        _meta_filtered_items(global_config, GLOBAL_CONFIG_EXCLUSIONS) if global_config else {}
    )

    aligned: list[dict[str, Any] | None] = [None] * len(setting_jobs)
    mismatches: dict[str, tuple[Any, Any]] = {}

    if not setting_jobs:
        return aligned, mismatches

    required = {"timestep", "learning_curve_mean", "learning_curve_std"}

    def _parse_sheets(sheet_map):
        out: list[tuple[str, dict[str, str], np.ndarray, np.ndarray, np.ndarray, Any, Any]] = []
        for sheet_name in sorted(sheet_map.keys()):
            df = sheet_map[sheet_name]
            if not required.issubset(set(df.columns)):
                continue
            try:
                ts = df["timestep"].values.astype(np.int32)
                lc = df["learning_curve_mean"].values.astype(np.float32)
                lc_s = df["learning_curve_std"].values.astype(np.float32)
            except Exception:
                continue
            sheet_hp: dict[str, str] = {}
            for col in df.columns:
                col_name = str(col)
                if col_name in required or col_name == "curve_label" or col_name.startswith("rep_"):
                    continue
                sheet_hp[col_name] = _parse_sheet_value(df[col].iloc[0])
            rep_cols_sorted = sorted(
                (c for c in df.columns if str(c).startswith("rep_")),
                key=lambda s: int(str(s).split("_", 1)[1]) if str(s).split("_", 1)[1].isdigit() else 10**9,
            )
            raw_returns = None
            if rep_cols_sorted:
                try:
                    raw_returns = np.asarray(
                        [df[c].values.astype(np.float32) for c in rep_cols_sorted],
                        dtype=np.float32,
                    )
                except Exception:
                    raw_returns = None
            out.append((sheet_name, sheet_hp, ts, lc, lc_s, df, raw_returns))
        return out

    header_row = 1 if formatted_sheets else 0
    try:
        sheets = pd.read_excel(filepath, sheet_name=None, engine="openpyxl", header=header_row)
    except Exception as exc:
        raise ValueError(f"Failed to read Excel file '{filepath}': {exc}") from exc

    parsed = _parse_sheets(sheets)
    if not parsed:
        try:
            sheets = pd.read_excel(
                filepath, sheet_name=None, engine="openpyxl",
                header=(0 if formatted_sheets else 1),
            )
        except Exception:
            return aligned, mismatches
        parsed = _parse_sheets(sheets)
    if not parsed:
        return aligned, mismatches

    legend: dict[str, tuple[str, bool]] = _resolve_legend_flags(
        algo_config or {}, warn_on_suppression=False
    )
    basename = os.path.basename(filepath)
    algo_prefix = os.path.splitext(basename)[0].upper()

    used: set[int] = set()
    for job_idx, job in enumerate(setting_jobs):
        job_hp = dict(job["hyperparams"])
        for gc_key, gc_text in global_filtered.items():
            job_hp.setdefault(gc_key, gc_text)

        try:
            running_n_ts = int(float(_sheet_hp_text(job_hp.get(GLOBAL_CONFIG_NTIMESTEPS_KEY))))
        except (TypeError, ValueError):
            running_n_ts = None

        for sheet_idx, (_sn, sheet_hp, ts, lc, lc_s, df, raw_returns) in enumerate(parsed):
            if sheet_idx in used:
                continue

            ok = True
            for key, job_val in job_hp.items():
                if key in ALGO_CONFIG_EXCLUSIONS or key in GLOBAL_CONFIG_EXCLUSIONS:
                    continue
                job_text = _sheet_hp_text(job_val)
                if key in sheet_hp:
                    sheet_text = sheet_hp[key]
                    if key == GLOBAL_CONFIG_NTIMESTEPS_KEY:
                        try:
                            if int(float(sheet_text)) < int(float(job_text)):
                                mismatches.setdefault(
                                    key, (sheet_text, f"{job_text} (need disk >= running)")
                                )
                                ok = False
                                break
                            continue
                        except (TypeError, ValueError):
                            pass
                    if sheet_text != job_text:
                        mismatches.setdefault(key, (sheet_text, job_text))
                        ok = False
                        break
                else:
                    if key == "eval_with_env_episode_trials":
                        if str(job_text).lower() == "true":
                            mismatches.setdefault(key, ("<missing_column>", job_text))
                            ok = False
                            break
                    elif key == "n_eval_episodes":
                        eet = _sheet_hp_text(job_hp.get("eval_with_env_episode_trials", False))
                        if str(eet).lower() == "true":
                            mismatches.setdefault(key, ("<missing_column>", job_text))
                            ok = False
                            break
                    # else: missing column tolerated for back-compat.
            if not ok:
                continue

            ts_out, lc_out, lc_s_out = ts, lc, lc_s
            raw_out = raw_returns
            if running_n_ts is not None:
                mask = ts_out <= running_n_ts
                ts_out = ts_out[mask]
                lc_out = lc_out[mask]
                lc_s_out = lc_s_out[mask]
                if raw_out is not None and raw_out.ndim == 2 and raw_out.shape[1] == len(mask):
                    raw_out = raw_out[:, mask]

            label_parts = [algo_prefix]
            for legend_key, (legend_label, show) in legend.items():
                if not show:
                    continue
                if legend_key in df.columns:
                    val = _parse_sheet_value(df[legend_key].iloc[0])
                elif legend_key in algo_config:
                    val = algo_config[legend_key]
                else:
                    continue
                legend_label = _format_legend_label(legend_label)
                if isinstance(val, bool):
                    label_parts.append(f"{legend_label}{val}")
                else:
                    label_parts.append(f"{legend_label}{_format_legend_value(val)}")

            aligned[job_idx] = {
                "learning_curve": lc_out,
                "learning_curve_std": lc_s_out,
                "timesteps": ts_out,
                "curve_label": ", ".join(label_parts),
                "raw_returns": raw_out,
            }
            used.add(sheet_idx)
            break

    return aligned, mismatches


def _hp_value_matches(cfg_key, cfg_val, sheet_val):
    """Check whether a sheet's HP value is compatible with the config value."""
    sheet_text = _sheet_hp_text(sheet_val)
    return sheet_text in _sheet_hp_text_candidates(cfg_val)


def _values_equal(a, b):
    """Compare two scalar values with float tolerance."""
    try:
        fa, fb = float(a), float(b)
        return abs(fa - fb) < 1e-7
    except (TypeError, ValueError):
        pass
    # Handle booleans explicitly
    if isinstance(a, (bool, np.bool_)) or isinstance(b, (bool, np.bool_)):
        return bool(a) == bool(b)
    return str(a) == str(b)


def _load_all_excel_curves(
    data_sheets_dir,
    algo_configs=None,
    formatted_sheets: bool = False,
    *,
    global_config: dict[str, Any] | None = None,
    original_algo_configs: dict[str, dict[str, Any]] | None = None,
):
    """Load all .xlsx files from data_sheets_dir (non-recursive).

    algo_configs: dict mapping algo name (e.g. "REINFORCE", "DQN") to its config dict.
    Each file's algo is inferred from the filename stem, so only files named
    like 'REINFORCE.xlsx' or 'DQN.xlsx' are matched.
    Sheets are filtered by HP value matching and labels are built using legend_parameters,
    both delegated to _load_results_from_excel. When 'global_config' is supplied
    each sheet is also validated against the filtered global_config keys.

    Returns a list of dicts: [{curve_label, learning_curve, learning_curve_std, timesteps, source_file}, ...]
    """
    all_curves = []
    pattern = os.path.join(data_sheets_dir, "*.xlsx")
    for filepath in sorted(glob.glob(pattern)):
        basename = os.path.basename(filepath)
        algo_prefix = os.path.splitext(basename)[0].upper()
        algo_config = (algo_configs or {}).get(algo_prefix)
        original_algo_config = (original_algo_configs or {}).get(algo_prefix)
        try:
            results, _ = _load_results_from_excel(
                filepath,
                algo_config,
                formatted_sheets=formatted_sheets,
                global_config=global_config,
                original_algo_config=original_algo_config,
            )
        except Exception:
            continue
        for entry in results:
            entry["source_file"] = basename
            all_curves.append(entry)
    return all_curves


# ── Per-sheet config validation ───────────────────────────────────────────────

# Keys that must NOT participate in per-sheet matching. Mirrors the
# user-supplied exclusion list (see Experiment.py global_config / algo_config).
GLOBAL_CONFIG_EXCLUSIONS = frozenset({
    "UNUSED_CPU_CORES",
    "benchmark_curve",
    "benchmark_name",
    "plot_smoothing_window",
    "curve_confidence_interval",
    "curve_shaded_area_opacity",
    "curve_plot",
    "show_curve_plots",
    "animation_plot",
    "use_existing_disk_data",
    "use_existing_disk_trained_networks",
    "format_sheets",
    "formatted_sheets",
    "separate_algorithm_plots",
    "Environment",
    "baseline_model",
    "n_use_trained_model",
    "action_selection_method",
    "trained_model_reseed_seed",
})

ALGO_CONFIG_EXCLUSIONS = frozenset({
    "legend_parameters",
    "nn_include_hp_in_legend",
    "nn_include_lr_in_legend",
})

# n_timesteps is matched with >= rather than equality (workbook may contain
# longer training curves than the running project requires).
GLOBAL_CONFIG_NTIMESTEPS_KEY = "n_timesteps"


def _meta_value_text(value: Any) -> str:
    """Stable string form for a config value persisted to a sheet column."""
    return _value_to_text(value)


def _meta_filtered_items(config: dict[str, Any] | None, exclusions: frozenset) -> dict[str, str]:
    """Return the '{key: text}' mapping for 'config' after dropping excluded keys.

    Values are normalized to a stable string so writing and reading produce
    identical comparable representations.
    """
    if not config:
        return {}
    items: dict[str, str] = {}
    for key, value in config.items():
        if str(key) in exclusions:
            continue
        items[str(key)] = _meta_value_text(value)
    return items


# ── Per-algorithm filename builder ────────────────────────────────────────────

def _build_algo_filename(algo_name, algo_config, n_repetitions, n_timesteps, eval_interval):
    """Build the workbook filename stem for a single algorithm."""
    return build_algorithm_filename(algo_name)


# ── Legend parameter helpers ──────────────────────────────────────────────────

LegendEntry = Tuple[str, bool]


def _normalize_legend_entry(param_name: str, entry: Any) -> LegendEntry:
    """Normalize a legend entry to a (label, show_flag) pair."""
    if isinstance(entry, (list, tuple)) and len(entry) == 2:
        label, show = entry
    else:
        label, show = param_name, entry
    return str(label), bool(show)


def _resolve_legend_flags(cfg: dict[str, Any], *, warn_on_suppression: bool = True) -> dict[str, LegendEntry]:
    """Resolve which parameters should appear in the legend.

    Uses 'legend_parameters' dict if present. Falls back to (and is
    overridden by) the legacy 'nn_include_hp_in_legend' /
    'nn_include_lr_in_legend' flags for backward compatibility.

    When DQN epsilon decay is disabled ('epsilon_decay_interval == 0'),
    suppress the dependent epsilon-decay fields from the legend so the
    DQN labels match the shared REINFORCE/AC/A2C path. If
    'warn_on_suppression' is True, print one warning per suppressed field.

    Returns a dict {param_name: (display_label, bool)}.
    """
    legend: dict[str, LegendEntry] = {}
    raw_legend_parameters = cfg.get("legend_parameters", {})
    if isinstance(raw_legend_parameters, dict):
        for param_name, entry in raw_legend_parameters.items():
            legend[param_name] = _normalize_legend_entry(param_name, entry)

    # Legacy overrides
    if "nn_include_lr_in_legend" in cfg:
        lr_key = "learning_rate" if "learning_rate" in cfg else "actor_lr"
        current_lr_entry = legend.get(lr_key)
        if current_lr_entry is not None:
            current_lr_label = current_lr_entry[0]
        else:
            current_lr_label = lr_key
        legend[lr_key] = (current_lr_label, bool(cfg["nn_include_lr_in_legend"]))
    if "nn_include_hp_in_legend" in cfg:
        hp_flag = bool(cfg["nn_include_hp_in_legend"])
        nn_key = "nn_hidden_layer_widths" if "nn_hidden_layer_widths" in cfg else "actor_hidden_nn"
        current_nn_entry = legend.get(nn_key)
        if current_nn_entry is not None:
            current_nn_label = current_nn_entry[0]
        else:
            current_nn_label = nn_key
        legend[nn_key] = (current_nn_label, hp_flag)
        current_gamma_entry = legend.get("gamma")
        if current_gamma_entry is not None:
            current_gamma_label = current_gamma_entry[0]
        else:
            current_gamma_label = "gamma"
        legend["gamma"] = (current_gamma_label, hp_flag)

    if cfg.get("exploration_method", "egreedy") == "egreedy":
        try:
            epsilon_decay_interval = int(cfg.get("epsilon_decay_interval", 1))
        except (TypeError, ValueError):
            epsilon_decay_interval = 1
        if epsilon_decay_interval == 0:
            for key in ("epsilon_start", "epsilon_end", "epsilon_decay"):
                entry = legend.get(key)
                if entry is None or not entry[1]:
                    continue
                if warn_on_suppression:
                    display_label = entry[0] if str(entry[0]).strip() else key
                    print(
                        f"[DQN] Warning (plot): '{display_label}' legend entry suppressed because "
                        "epsilon_decay_interval=0 disables epsilon decay trials."
                    )
                legend[key] = (entry[0], False)

    return legend


def _fmt_legend(v: Any) -> str:
    """Format numbers for legend display.

    Uses scientific notation when |v| < 0.001 or |v| >= 1000.
    """
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, (int, float, np.integer, np.floating)):
        f = float(v)
        if f != 0.0 and (abs(f) < 1e-3 or abs(f) >= 1e3):
            mantissa, exp = f"{f:.3e}".split("e")
            mantissa = mantissa.rstrip("0").rstrip(".")
            return f"{mantissa}e{int(exp)}"
        if isinstance(v, (float, np.floating)):
            return f"{f:.3g}"
        return str(v)
    return str(v)


def _build_legend_parts(legend: dict[str, LegendEntry], cfg: dict[str, Any]) -> list[str]:
    """Build a list of 'label=value' strings for every legend-enabled parameter."""
    parts: list[str] = []
    for key, (label, show) in legend.items():
        if not show:
            continue
        val = cfg.get(key)
        if val is None:
            continue

        label = _format_legend_label(label)

        if isinstance(val, (bool, np.bool_)):
            parts.append(f"{label}{bool(val)}")
        else:
            parts.append(f"{label}{_format_legend_value(val)}")
    return parts


# ── Build setting jobs per algorithm ──────────────────────────────────────────

def _as_list(value):
    """Wrap a scalar in a single-element list; pass through lists/tuples/arrays."""
    if isinstance(value, (list, tuple)):
        return list(value)
    if isinstance(value, np.ndarray):
        return value.tolist()
    return [value]


def _parse_pg_config(cfg):
    """Parse policy-gradient config into sweepable arrays (with backward compat)."""
    gammas = np.atleast_1d(np.asarray(cfg.get("gamma", np.array([0.99])), dtype=np.float32))
    learning_rates = np.atleast_1d(np.asarray(cfg.get("actor_lr", np.array([0.001])), dtype=np.float32))

    raw_nn = cfg.get("actor_hidden_nn", [[32, 32]])
    if isinstance(raw_nn, np.ndarray) and raw_nn.ndim == 1:
        nn_architectures = [raw_nn]  # backward compat: np.array([32,32])
    elif isinstance(raw_nn, list) and len(raw_nn) > 0 and not isinstance(raw_nn[0], (list, np.ndarray)):
        nn_architectures = [np.asarray(raw_nn, dtype=np.int32)]  # backward compat: flat list [32,32]
    else:
        nn_architectures = [np.asarray(arch, dtype=np.int32) for arch in raw_nn]

    legend = _resolve_legend_flags(cfg)
    return gammas, learning_rates, nn_architectures, legend


def _build_reinforce_jobs(
    *,
    algo_config,
    n_repetitions,
    n_timesteps,
    eval_interval,
    max_train_episode_length,
    max_eval_episode_length,
    base_seed,
    eval_with_env_episode_trials: bool,
    n_eval_episodes: int,
):
    """Build REINFORCE training jobs."""
    cfg = algo_config
    gammas, learning_rates, nn_architectures, legend = _parse_pg_config(cfg)

    setting_jobs = []
    for gamma_val in gammas:
        gamma_val = float(gamma_val)
        for nn_arch in nn_architectures:
            nn_arch = np.asarray(nn_arch, dtype=np.int32)
            for lr_val in learning_rates:
                lr_val = float(lr_val)

                legend_cfg = {
                    **cfg,
                    "gamma": gamma_val,
                    "actor_lr": lr_val,
                    "actor_hidden_nn": nn_arch,
                }
                label_parts = ["REINFORCE"] + _build_legend_parts(legend, legend_cfg)
                curve_label = ", ".join(label_parts)

                setting_jobs.append({
                    "curve_label": curve_label,
                    "method": "REINFORCE",
                    "kwargs": dict(
                        method="REINFORCE",
                        n_repetitions=n_repetitions,
                        n_timesteps=n_timesteps,
                        eval_interval=eval_interval,
                        max_train_episode_length=max_train_episode_length,
                        max_eval_episode_length=max_eval_episode_length,
                        actor_lr=lr_val,
                        gamma=gamma_val,
                        actor_hidden_nn=nn_arch,
                        base_seed=base_seed,
                        eval_with_env_episode_trials=eval_with_env_episode_trials,
                        n_eval_episodes=n_eval_episodes,
                        plot_smoothing_window=1,
                    ),
                    "hyperparams": {
                        "n_repetitions": n_repetitions,
                        "n_timesteps": n_timesteps,
                        "eval_interval": eval_interval,
                        "max_train_episode_length": max_train_episode_length,
                        "max_eval_episode_length": max_eval_episode_length,
                        "actor_lr": lr_val,
                        "gamma": gamma_val,
                        "actor_hidden_nn": str(nn_arch.tolist()),
                        "eval_with_env_episode_trials": eval_with_env_episode_trials,
                        "n_eval_episodes": n_eval_episodes,
                    },
                })
    return setting_jobs


def _build_ac_jobs(
    *,
    algo_config,
    n_repetitions,
    n_timesteps,
    eval_interval,
    max_train_episode_length,
    max_eval_episode_length,
    base_seed,
    eval_with_env_episode_trials: bool,
    n_eval_episodes: int,
):
    cfg = algo_config
    gammas, learning_rates, nn_architectures, legend = _parse_pg_config(cfg)
    critic_learning_rates = np.atleast_1d(np.asarray(cfg.get("critic_lr", np.array([0.001])), dtype=np.float32))
    raw_critic_nn = cfg.get("critic_hidden_nn", [[64, 64]])
    if isinstance(raw_critic_nn, np.ndarray) and raw_critic_nn.ndim == 1:
        critic_architectures = [raw_critic_nn]
    elif isinstance(raw_critic_nn, list) and len(raw_critic_nn) > 0 and not isinstance(raw_critic_nn[0], (list, np.ndarray)):
        critic_architectures = [np.asarray(raw_critic_nn, dtype=np.int32)]
    else:
        critic_architectures = [np.asarray(arch, dtype=np.int32) for arch in raw_critic_nn]

    setting_jobs = []
    for gamma_val in gammas:
        gamma_val = float(gamma_val)
        for actor_nn in nn_architectures:
            actor_nn = np.asarray(actor_nn, dtype=np.int32)
            for actor_lr_val in learning_rates:
                actor_lr_val = float(actor_lr_val)
                for critic_nn in critic_architectures:
                    critic_nn = np.asarray(critic_nn, dtype=np.int32)
                    for critic_lr_val in critic_learning_rates:
                        critic_lr_val = float(critic_lr_val)
                        iter_cfg = {
                            **cfg,
                            "gamma": gamma_val,
                            "actor_lr": actor_lr_val,
                            "actor_hidden_nn": actor_nn,
                            "critic_lr": critic_lr_val,
                            "critic_hidden_nn": critic_nn,
                        }
                        label_parts = ["AC"] + _build_legend_parts(legend, iter_cfg)
                        curve_label = ", ".join(label_parts)
                        setting_jobs.append({
                            "curve_label": curve_label,
                            "method": "ac",
                            "kwargs": dict(
                                method="ac",
                                n_repetitions=n_repetitions,
                                n_timesteps=n_timesteps,
                                eval_interval=eval_interval,
                                max_train_episode_length=max_train_episode_length,
                                max_eval_episode_length=max_eval_episode_length,
                                actor_lr=actor_lr_val,
                                critic_lr=critic_lr_val,
                                gamma=gamma_val,
                                actor_hidden_nn=actor_nn,
                                critic_hidden_nn=critic_nn,
                                base_seed=base_seed,
                                eval_with_env_episode_trials=eval_with_env_episode_trials,
                                n_eval_episodes=n_eval_episodes,
                                plot_smoothing_window=1,
                            ),
                            "hyperparams": {
                                "n_repetitions": n_repetitions,
                                "n_timesteps": n_timesteps,
                                "eval_interval": eval_interval,
                                "max_train_episode_length": max_train_episode_length,
                                "max_eval_episode_length": max_eval_episode_length,
                                "actor_lr": actor_lr_val,
                                "critic_lr": critic_lr_val,
                                "gamma": gamma_val,
                                "actor_hidden_nn": str(actor_nn.tolist()),
                                "critic_hidden_nn": str(critic_nn.tolist()),
                                "eval_with_env_episode_trials": eval_with_env_episode_trials,
                                "n_eval_episodes": n_eval_episodes,
                            },
                        })
    return setting_jobs


def _build_a2c_jobs(
    *,
    algo_config,
    n_repetitions,
    n_timesteps,
    eval_interval,
    max_train_episode_length,
    max_eval_episode_length,
    base_seed,
    eval_with_env_episode_trials: bool,
    n_eval_episodes: int,
):
    cfg = algo_config
    gammas, actor_learning_rates, actor_architectures, legend = _parse_pg_config(cfg)
    critic_learning_rates = np.atleast_1d(np.asarray(cfg.get("critic_lr", np.array([0.001])), dtype=np.float32))
    raw_critic_nn = cfg.get("critic_hidden_nn", [[64, 64]])
    if isinstance(raw_critic_nn, np.ndarray) and raw_critic_nn.ndim == 1:
        critic_architectures = [raw_critic_nn]
    elif isinstance(raw_critic_nn, list) and len(raw_critic_nn) > 0 and not isinstance(raw_critic_nn[0], (list, np.ndarray)):
        critic_architectures = [np.asarray(raw_critic_nn, dtype=np.int32)]
    else:
        critic_architectures = [np.asarray(arch, dtype=np.int32) for arch in raw_critic_nn]
    TN_steps = np.atleast_1d(np.asarray(cfg.get("TN_step", np.array([10])), dtype=np.int32))

    setting_jobs = []
    for gamma_val in gammas:
        gamma_val = float(gamma_val)
        for actor_nn in actor_architectures:
            actor_nn = np.asarray(actor_nn, dtype=np.int32)
            for actor_lr_val in actor_learning_rates:
                actor_lr_val = float(actor_lr_val)
                for critic_nn in critic_architectures:
                    critic_nn = np.asarray(critic_nn, dtype=np.int32)
                    for critic_lr_val in critic_learning_rates:
                        critic_lr_val = float(critic_lr_val)
                        for tn_step in TN_steps:
                            tn_step = int(tn_step)
                            iter_cfg = {
                                **cfg,
                                "gamma": gamma_val,
                                "actor_lr": actor_lr_val,
                                "actor_hidden_nn": actor_nn,
                                "critic_lr": critic_lr_val,
                                "critic_hidden_nn": critic_nn,
                                "TN_step": tn_step,
                            }
                            label_parts = ["A2C"] + _build_legend_parts(legend, iter_cfg)
                            curve_label = ", ".join(label_parts)
                            setting_jobs.append({
                                "curve_label": curve_label,
                                "method": "a2c",
                                "kwargs": dict(
                                    method="a2c",
                                    n_repetitions=n_repetitions,
                                    n_timesteps=n_timesteps,
                                    eval_interval=eval_interval,
                                    max_train_episode_length=max_train_episode_length,
                                    max_eval_episode_length=max_eval_episode_length,
                                    actor_lr=actor_lr_val,
                                    critic_lr=critic_lr_val,
                                    gamma=gamma_val,
                                    actor_hidden_nn=actor_nn,
                                    critic_hidden_nn=critic_nn,
                                    TN_step=tn_step,
                                    base_seed=base_seed,
                                    eval_with_env_episode_trials=eval_with_env_episode_trials,
                                    n_eval_episodes=n_eval_episodes,
                                    plot_smoothing_window=1,
                                ),
                                "hyperparams": {
                                    "n_repetitions": n_repetitions,
                                    "n_timesteps": n_timesteps,
                                    "eval_interval": eval_interval,
                                    "max_train_episode_length": max_train_episode_length,
                                    "max_eval_episode_length": max_eval_episode_length,
                                    "actor_lr": actor_lr_val,
                                    "critic_lr": critic_lr_val,
                                    "gamma": gamma_val,
                                    "actor_hidden_nn": str(actor_nn.tolist()),
                                    "critic_hidden_nn": str(critic_nn.tolist()),
                                    "TN_step": tn_step,
                                    "eval_with_env_episode_trials": eval_with_env_episode_trials,
                                    "n_eval_episodes": n_eval_episodes,
                                },
                            })
    return setting_jobs


def _build_ppo_jobs(
    *,
    algo_config,
    n_repetitions,
    n_timesteps,
    eval_interval,
    max_train_episode_length,
    max_eval_episode_length,
    base_seed,
    eval_with_env_episode_trials: bool,
    n_eval_episodes: int,
):
    cfg = algo_config
    gammas, actor_learning_rates, actor_architectures, legend = _parse_pg_config(cfg)
    critic_learning_rates = np.atleast_1d(np.asarray(cfg.get("critic_lr", np.array([0.001])), dtype=np.float32))
    raw_critic_nn = cfg.get("critic_hidden_nn", [[64, 64]])
    if isinstance(raw_critic_nn, np.ndarray) and raw_critic_nn.ndim == 1:
        critic_architectures = [raw_critic_nn]
    elif isinstance(raw_critic_nn, list) and len(raw_critic_nn) > 0 and not isinstance(raw_critic_nn[0], (list, np.ndarray)):
        critic_architectures = [np.asarray(raw_critic_nn, dtype=np.int32)]
    else:
        critic_architectures = [np.asarray(arch, dtype=np.int32) for arch in raw_critic_nn]

    gae_lambdas = np.atleast_1d(np.asarray(cfg.get("gae_lambda", np.array([0.95])), dtype=np.float32))
    clip_eps_list = np.atleast_1d(np.asarray(cfg.get("clip_eps", np.array([0.2])), dtype=np.float32))
    n_epochs_list = np.atleast_1d(np.asarray(cfg.get("n_epochs", np.array([10])), dtype=np.int32))
    rollout_steps_list = np.atleast_1d(np.asarray(cfg.get("rollout_steps", np.array([2048])), dtype=np.int32))

    setting_jobs = []
    for gamma_val in gammas:
        gamma_val = float(gamma_val)
        for actor_nn in actor_architectures:
            actor_nn = np.asarray(actor_nn, dtype=np.int32)
            for actor_lr_val in actor_learning_rates:
                actor_lr_val = float(actor_lr_val)
                for critic_nn in critic_architectures:
                    critic_nn = np.asarray(critic_nn, dtype=np.int32)
                    for critic_lr_val in critic_learning_rates:
                        critic_lr_val = float(critic_lr_val)
                        for gae_lambda_val in gae_lambdas:
                            gae_lambda_val = float(gae_lambda_val)
                            for clip_eps_val in clip_eps_list:
                                clip_eps_val = float(clip_eps_val)
                                for n_epochs_val in n_epochs_list:
                                    n_epochs_val = int(n_epochs_val)
                                    for rollout_steps_val in rollout_steps_list:
                                        rollout_steps_val = int(rollout_steps_val)
                                        iter_cfg = {
                                            **cfg,
                                            "gamma": gamma_val,
                                            "actor_lr": actor_lr_val,
                                            "actor_hidden_nn": actor_nn,
                                            "critic_lr": critic_lr_val,
                                            "critic_hidden_nn": critic_nn,
                                            "gae_lambda": gae_lambda_val,
                                            "clip_eps": clip_eps_val,
                                            "n_epochs": n_epochs_val,
                                            "rollout_steps": rollout_steps_val,
                                        }
                                        label_parts = ["PPO"] + _build_legend_parts(legend, iter_cfg)
                                        curve_label = ", ".join(label_parts)
                                        setting_jobs.append({
                                            "curve_label": curve_label,
                                            "method": "ppo",
                                            "kwargs": dict(
                                                method="ppo",
                                                n_repetitions=n_repetitions,
                                                n_timesteps=n_timesteps,
                                                eval_interval=eval_interval,
                                                max_train_episode_length=max_train_episode_length,
                                                max_eval_episode_length=max_eval_episode_length,
                                                actor_lr=actor_lr_val,
                                                critic_lr=critic_lr_val,
                                                gamma=gamma_val,
                                                actor_hidden_nn=actor_nn,
                                                critic_hidden_nn=critic_nn,
                                                gae_lambda=gae_lambda_val,
                                                clip_eps=clip_eps_val,
                                                n_epochs=n_epochs_val,
                                                rollout_steps=rollout_steps_val,
                                                base_seed=base_seed,
                                                eval_with_env_episode_trials=eval_with_env_episode_trials,
                                                n_eval_episodes=n_eval_episodes,
                                                plot_smoothing_window=1,
                                            ),
                                            "hyperparams": {
                                                "n_repetitions": n_repetitions,
                                                "n_timesteps": n_timesteps,
                                                "eval_interval": eval_interval,
                                                "max_train_episode_length": max_train_episode_length,
                                                "max_eval_episode_length": max_eval_episode_length,
                                                "actor_lr": actor_lr_val,
                                                "critic_lr": critic_lr_val,
                                                "gamma": gamma_val,
                                                "actor_hidden_nn": str(actor_nn.tolist()),
                                                "critic_hidden_nn": str(critic_nn.tolist()),
                                                "gae_lambda": gae_lambda_val,
                                                "clip_eps": clip_eps_val,
                                                "n_epochs": n_epochs_val,
                                                "rollout_steps": rollout_steps_val,
                                                "eval_with_env_episode_trials": eval_with_env_episode_trials,
                                                "n_eval_episodes": n_eval_episodes,
                                            },
                                        })
    return setting_jobs


def _build_dqn_jobs(*, dqn_config, n_repetitions, n_timesteps, eval_interval,
                    max_train_episode_length, max_eval_episode_length, base_seed,
                    eval_with_env_episode_trials: bool, n_eval_episodes: int):
    """Build setting_jobs for DQN using assignment2_repo infrastructure."""
    cfg = dqn_config
    learning_rates = np.atleast_1d(np.asarray(cfg.get("learning_rate", np.array([0.001])), dtype=np.float32))
    raw_nn = cfg.get("nn_hidden_layer_widths", [[128, 128]])
    if isinstance(raw_nn, np.ndarray) and raw_nn.ndim == 1:
        nn_architectures = [raw_nn]
    elif isinstance(raw_nn, list) and len(raw_nn) > 0 and not isinstance(raw_nn[0], (list, np.ndarray)):
        nn_architectures = [np.asarray(raw_nn, dtype=np.int32)]
    else:
        nn_architectures = [np.asarray(arch, dtype=np.int32) for arch in raw_nn]
    gamma = float(cfg.get("gamma", 0.99))
    exploration_method = cfg.get("exploration_method", "egreedy")
    epsilons = np.atleast_1d(np.asarray(cfg.get("epsilons", np.array([0.05])), dtype=np.float32))
    epsilon_start = float(cfg.get("epsilon_start", 1.0))
    epsilon_end = float(cfg.get("epsilon_end", 0.02))
    epsilon_decay = float(cfg.get("epsilon_decay", 0.9995))
    epsilon_decay_interval = int(cfg.get("epsilon_decay_interval", 0))
    softmax_temps = np.atleast_1d(np.asarray(cfg.get("softmax_temps", np.array([1.0, 0.5, 0.1])), dtype=np.float32))

    TN_active = np.atleast_1d(np.asarray(cfg.get("TN_active", np.array([True]))))
    TN_step = np.atleast_1d(np.asarray(cfg.get("TN_step", np.array([100])), dtype=np.int32))
    ER_active = np.atleast_1d(np.asarray(cfg.get("ER_active", np.array([True]))))
    ER_replay_buffer_size = int(cfg.get("ER_replay_buffer_size", 80000))
    ER_batch_size = np.atleast_1d(np.asarray(cfg.get("ER_batch_size", np.array([64])), dtype=np.int32))
    ER_min_replay_size = int(cfg.get("ER_min_replay_size", 2000))
    ER_sample_train_frequency = np.atleast_1d(np.asarray(cfg.get("ER_sample_train_frequency", np.array([1])), dtype=np.int32))
    ER_replay_ratio = float(cfg.get("ER_replay_ratio", 1.0))

    legend = _resolve_legend_flags(cfg)

    epsilon_decay_enabled = epsilon_decay_interval > 0
    skip_decay_trials = exploration_method == "egreedy" and not epsilon_decay_enabled

    eval_with_env_episode_trials = bool(eval_with_env_episode_trials)
    n_eval_episodes = int(n_eval_episodes)

    setting_jobs = []
    for nn_arch in nn_architectures:
        nn_arch = np.asarray(nn_arch, dtype=np.int32)
        for tn_active in TN_active:
            tn_active_bool = bool(tn_active)
            effective_tn_steps = TN_step if tn_active_bool else np.array([1])
            for er_active_val in ER_active:
                er_active_bool = bool(er_active_val)
                for er_bs in ER_batch_size:
                    er_bs = int(er_bs)
                    for er_freq in ER_sample_train_frequency:
                        er_freq = int(er_freq)
                        for lr_val in learning_rates:
                            lr_val = float(lr_val)
                            for tn_step in effective_tn_steps:
                                tn_step = int(tn_step)
                                iter_cfg = {**cfg, "nn_hidden_layer_widths": nn_arch,
                                            "learning_rate": lr_val, "gamma": gamma}
                                label_parts = ["DQN"]
                                label_parts.extend(_build_legend_parts(legend, iter_cfg))
                                run_label_prefix = ", ".join(label_parts)

                                er_kwargs = dict(
                                    er_active=er_active_bool,
                                    er_replay_buffer_size=ER_replay_buffer_size,
                                    er_batch_size=er_bs,
                                    er_min_replay_size=ER_min_replay_size,
                                    er_sample_train_frequency=er_freq,
                                    er_replay_ratio=ER_replay_ratio,
                                )
                                common = dict(
                                    n_timesteps=n_timesteps,
                                    max_episode_length=max_train_episode_length,
                                    max_eval_episode_length=max_eval_episode_length,
                                    eval_with_env_episode_trials=eval_with_env_episode_trials,
                                    n_eval_episodes=n_eval_episodes,
                                    learning_rate=lr_val,
                                    gamma=gamma,
                                    nn_hidden_layer_widths=nn_arch,
                                    TN_step=tn_step,
                                    target_network_active=tn_active_bool,
                                    eval_interval=eval_interval,
                                    base_seed=base_seed,
                                    exploration_method=exploration_method,
                                    **er_kwargs,
                                )
                                base_hp = {
                                    "n_repetitions": n_repetitions,
                                    "n_timesteps": n_timesteps,
                                    "eval_interval": eval_interval,
                                    "max_episode_length": max_train_episode_length,
                                    "max_eval_episode_length": max_eval_episode_length,
                                    "eval_with_env_episode_trials": eval_with_env_episode_trials,
                                    "n_eval_episodes": n_eval_episodes,
                                    "learning_rate": lr_val,
                                    "gamma": gamma,
                                    "nn_hidden_layer_widths": str(nn_arch.tolist()),
                                    "exploration_method": exploration_method,
                                    "TN_active": tn_active_bool,
                                    "TN_step": tn_step,
                                    "ER_active": er_active_bool,
                                    "ER_batch_size": er_bs,
                                    "ER_sample_train_frequency": er_freq,
                                }

                                # epsilon-decay trial
                                if not skip_decay_trials and exploration_method == "egreedy":
                                    setting_jobs.append({
                                        "curve_label": run_label_prefix,
                                        "method": "dqn",
                                        "kwargs": {
                                            **common,
                                            "n_repetitions": n_repetitions,
                                            "epsilon_start": epsilon_start,
                                            "epsilon_end": epsilon_end,
                                            "epsilon_decay": epsilon_decay,
                                            "epsilon_decay_interval": epsilon_decay_interval,
                                            "softmax_temp": None,
                                        },
                                        "hyperparams": {**base_hp,
                                            "epsilon_start": epsilon_start,
                                            "epsilon_end": epsilon_end,
                                        },
                                    })

                                # fixed-epsilon trials
                                if exploration_method == "egreedy":
                                    for eps in epsilons:
                                        setting_jobs.append({
                                            "curve_label": run_label_prefix,
                                            "method": "dqn",
                                            "kwargs": {
                                                **common,
                                                "n_repetitions": n_repetitions,
                                                "epsilon_start": float(eps),
                                                "epsilon_end": float(eps),
                                                "epsilon_decay": 1.0,
                                                "epsilon_decay_interval": 1,
                                                "softmax_temp": None,
                                            },
                                            "hyperparams": {**base_hp,
                                                "epsilon_start": float(eps),
                                                "epsilon_end": float(eps),
                                            },
                                        })
                                else:
                                    for temp in softmax_temps:
                                        setting_jobs.append({
                                            "curve_label": run_label_prefix,
                                            "method": "dqn",
                                            "kwargs": {
                                                **common,
                                                "n_repetitions": n_repetitions,
                                                "epsilon_start": None,
                                                "epsilon_end": None,
                                                "epsilon_decay": 1.0,
                                                "epsilon_decay_interval": 1,
                                                "softmax_temp": float(temp),
                                            },
                                            "hyperparams": {**base_hp,
                                                "softmax_temp": float(temp),
                                            },
                                        })
    return setting_jobs


# ── Cross-setting parallel execution engine ───────────────────────────────────────────

def _dqn_job_kwargs_to_trial_common(job_kwargs: dict[str, Any], eval_interval: int) -> dict[str, Any]:
    """Map a DQN job's kwargs dict to the run_dqn_trial_returns keyword arguments."""
    kw: dict[str, Any] = dict(job_kwargs)
    exploration_method = kw.get("exploration_method", "egreedy")
    trial_common: dict[str, Any] = dict(
        n_env_steps=int(kw["n_timesteps"]),
        max_episode_length=int(kw["max_episode_length"]),
        max_eval_episode_length=int(kw.get("max_eval_episode_length", kw["max_episode_length"])),
        eval_with_env_episode_trials=bool(kw.get("eval_with_env_episode_trials", False)),
        n_eval_episodes=int(kw.get("n_eval_episodes", 5)),
        learning_rate=float(kw["learning_rate"]),
        nn_hidden_layer_widths=np.asarray(kw["nn_hidden_layer_widths"], dtype=np.int32),
        discount_factor=float(kw["gamma"]),
        target_network_step=int(kw["TN_step"]),
        target_network_active=bool(kw["target_network_active"]),
        n_returns_interval=eval_interval,
        exploration_method="epsilon_greedy" if exploration_method == "egreedy" else exploration_method,
        epsilon_start=kw.get("epsilon_start"),
        epsilon_end=kw.get("epsilon_end"),
        epsilon_decay=float(kw.get("epsilon_decay", 1.0)),
        epsilon_decay_interval=int(kw.get("epsilon_decay_interval", 1)),
        softmax_temp=kw.get("softmax_temp"),
        er_active=bool(kw.get("er_active", False)),
        er_replay_buffer_size=int(kw.get("er_replay_buffer_size", 10000)),
        er_batch_size=int(kw.get("er_batch_size", 64)),
        er_min_replay_size=int(kw.get("er_min_replay_size", 100)),
        er_sample_train_frequency=int(kw.get("er_sample_train_frequency", 1)),
        er_replay_ratio=float(kw.get("er_replay_ratio", 1.0)),
        emit_trial_header=False,
    )
    return trial_common


def _run_dqn_one_rep(trial_common: dict[str, Any], run_seed: int, rep_index: int, n_repetitions: int,
                     shared_step_counter=None,
                     use_existing_disk_trained_networks: bool = False):
    """Run one DQN repetition. Pickle-safe for ProcessPoolExecutor.

    Mirrors the checkpointing behaviour of the policy-gradient algorithms:
    optionally pre-loads the Q-network from disk before training, and on the
    first repetition ('rep_index == 0') persists the trained Q-network to
    disk via :func:`Checkpointing.save_state_dict_overwrite`.
    """
    import torch
    from assignment2_repo.DQN import run_dqn_trial_returns
    from Checkpointing import (
        dqn_q_checkpoint_path,
        save_state_dict_overwrite,
    )

    ck = dqn_q_checkpoint_path(
        nn_hidden_layer_widths=trial_common["nn_hidden_layer_widths"],
    )

    pretrained_state_dict = None
    if use_existing_disk_trained_networks and os.path.isfile(ck.file_path):
        pretrained_state_dict = torch.load(ck.file_path, map_location="cpu")

    returns_arr, timesteps_arr, model = run_dqn_trial_returns(
        seed=run_seed,
        trial_run_index=rep_index + 1,
        total_trial_runs=n_repetitions,
        enable_progress_bar=False,
        shared_step_counter=shared_step_counter,
        pretrained_state_dict=pretrained_state_dict,
        return_model=True,
        **trial_common,
    )

    if rep_index == 0:
        save_state_dict_overwrite(
            model=model,
            checkpoint_path=ck.file_path,
        )

    return returns_arr, timesteps_arr


def _run_pending_parallel(pending_settings, n_repetitions, n_timesteps, eval_interval,
                          max_train_episode_length, max_eval_episode_length, base_seed,
                          use_existing_disk_trained_networks: bool,
    setting_results: list[
        tuple[np.ndarray, np.ndarray, np.ndarray] | tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray]
        | None
    ],
                          unused_cpu_cores: int = 0,
                          on_setting_complete=None,
                          poll_callback=None):
    """Run all pending (setting × rep) tasks in a single flat ProcessPoolExecutor.

    pending_settings: list of (global_idx, job)
    Fills setting_results[global_idx] = (lc_mean, lc_std, timesteps) for each entry.
    Progress bars mirror the current per-rep format; a (S{n}) suffix is added when
    more than one setting is running simultaneously.

    on_setting_complete: optional callable(global_idx) invoked from the main
        process as soon as every repetition of a single setting has finished and
        its aggregated entry has been written to setting_results.
    poll_callback: optional callable invoked once per polling tick (after sleep).
        Used by the caller to pump the matplotlib GUI event loop while
        per-algo plot windows are being shown non-blocking.
    """
    from Library import _run_single_repetition

    total_tasks = len(pending_settings) * n_repetitions
    multi = len(pending_settings) > 1

    with Manager() as mgr:
        step_counters = {}
        for sp, (_, job) in enumerate(pending_settings):
            for r in range(n_repetitions):
                step_counters[(sp, r)] = mgr.Value('i', 0)

        cpu_count = os.cpu_count() or 1
        if unused_cpu_cores is None:
            unused_cpu_cores = 0
        unused_cpu_cores = int(unused_cpu_cores)
        if unused_cpu_cores < 0:
            unused_cpu_cores = 0
        available_cpus = max(1, cpu_count - unused_cpu_cores)
        max_workers = min(total_tasks, available_cpus)
        with ProcessPoolExecutor(max_workers=max_workers) as executor:
            futures = {}
            for sp, (_, job) in enumerate(pending_settings):
                method = job["method"]
                kw: dict[str, Any] = dict(job["kwargs"])
                for r in range(n_repetitions):
                    run_seed = base_seed + r
                    if method == "dqn":
                        trial_common = _dqn_job_kwargs_to_trial_common(kw, eval_interval)
                        future = executor.submit(
                            _run_dqn_one_rep,
                            trial_common=trial_common,
                            run_seed=run_seed,
                            rep_index=r,
                            n_repetitions=n_repetitions,
                            shared_step_counter=step_counters[(sp, r)],
                            use_existing_disk_trained_networks=use_existing_disk_trained_networks,
                        )
                    else:
                            algo_extra_kwargs: dict[str, Any] = {}
                            # A2C-specific
                            if "TN_step" in kw:
                                algo_extra_kwargs["TN_step"] = int(kw["TN_step"])
                            # PPO-specific
                            for k in (
                                "gae_lambda",
                                "clip_eps",
                                "n_epochs",
                                "rollout_steps",
                            ):
                                if k in kw:
                                    algo_extra_kwargs[k] = kw[k]
                            future = executor.submit(
                                _run_single_repetition,
                                method=method,
                                actor_hidden_nn=kw["actor_hidden_nn"],
                                critic_hidden_nn=kw.get("critic_hidden_nn", np.array([64, 64])),
                                actor_lr=kw["actor_lr"],
                                critic_lr=kw.get("critic_lr", 0.001),
                                gamma=kw["gamma"],
                                max_train_episode_length=max_train_episode_length,
                                max_eval_episode_length=max_eval_episode_length,
                                n_timesteps=n_timesteps,
                                eval_interval=eval_interval,
                                run_seed=run_seed,
                                rep_index=r,
                                n_repetitions=n_repetitions,
                                enable_progress_bar=False,
                                shared_step_counter=step_counters[(sp, r)],
                                eval_with_env_episode_trials=bool(kw.get("eval_with_env_episode_trials", False)),
                                n_eval_episodes=int(kw.get("n_eval_episodes", 5)),
                                use_existing_disk_trained_networks=use_existing_disk_trained_networks,
                                **algo_extra_kwargs,
                            )
                    futures[future] = (sp, r)

            max_visible_bars = int(os.environ.get("RL_MAX_TQDM_BARS", "10"))

            # Build a stable task list based on which step_counters exist.
            task_keys: list[tuple[int, int]] = list(step_counters.keys())

            # Create merged/grouped bars so we never exceed max_visible_bars.
            # When a bar merges k reps, its tqdm total is multiplied by k and we
            # update by the sum of the merged reps' progress -> it fills slower.
            total_rep_tasks = len(task_keys)
            if total_rep_tasks == 0:
                pbars = {}
                rep_last: dict[tuple[int, int], int] = {}
                rep_to_group: dict[tuple[int, int], int] = {}
                groups: list[list[tuple[int, int]]] = []
            else:
                n_groups = max(1, min(max_visible_bars, total_rep_tasks))
                group_size = (total_rep_tasks + n_groups - 1) // n_groups

                groups = [
                    task_keys[i * group_size : min((i + 1) * group_size, total_rep_tasks)]
                    for i in range(n_groups)
                ]
                # Drop potential empty trailing groups
                groups = [g for g in groups if g]

                rep_to_group = {}
                for gi, group in enumerate(groups):
                    for key in group:
                        rep_to_group[key] = gi

                rep_last = {key: 0 for key in task_keys}

                pbars = {}
                for gi, group in enumerate(groups):
                    (sp0, _r0) = group[0]
                    m = pending_settings[sp0][1]["method"].upper()

                    same_sp_and_method = True
                    sp0_val = group[0][0]
                    for (sp, _r) in group:
                        if sp != sp0_val or pending_settings[sp][1]["method"] != pending_settings[sp0_val][1]["method"]:
                            same_sp_and_method = False
                            break

                    if same_sp_and_method:
                        rep_ids = sorted(r + 1 for (_sp, r) in group)
                        rep_list_str = ",".join(str(x) for x in rep_ids)
                        suffix = f" (S{sp0_val + 1})" if multi else ""
                        desc = f"{m} Rep {rep_list_str}/{n_repetitions}{suffix}"
                    else:
                        # Fallback: show which (S,Rep) are merged.
                        entries = ",".join(f"S{sp + 1}R{r + 1}" for (sp, r) in group)
                        desc = f"{m} Rep {entries} (merged)/{n_repetitions}"

                    pbars[gi] = tqdm(
                        total=n_timesteps * len(group),
                        desc=desc,
                        unit="step",
                        position=gi,
                        leave=True,
                        dynamic_ncols=True,
                        bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]",
                    )

            rep_results = {}
            done = set()
            aggregated_settings: set[int] = set()
            try:
                while len(done) < total_tasks:
                    # Update merged bars from shared step_counters.
                    for (sp, r), sc in step_counters.items():
                        pb_gi = rep_to_group.get((sp, r))
                        if pb_gi is None:
                            continue
                        cur = sc.value
                        last = rep_last[(sp, r)]
                        delta = cur - last
                        if delta > 0:
                            pbars[pb_gi].update(delta)
                            rep_last[(sp, r)] = cur

                    for f in list(futures):
                        if f not in done and f.done():
                            done.add(f)
                            sp, r = futures[f]
                            try:
                                rep_results[(sp, r)] = f.result()
                            except Exception as exc:
                                if use_existing_disk_trained_networks:
                                    print(
                                        "\n[Trial failure detected with use_existing_disk_trained_networks=True]\n"
                                        "Suggestion: verify that the model architecture hyperparameters\n"
                                        "(actor_hidden_nn / critic_hidden_nn / TN_step, etc.) match the\n"
                                        "saved checkpoint signature in Checkpoints/. If they differ,\n"
                                        "set use_existing_disk_trained_networks=False to resume training.\n"
                                        f"Error: {exc}\n"
                                    )
                                    try:
                                        input("Press Enter to re-raise the error (no fallback will be applied) ...")
                                    except Exception:
                                        pass
                                raise
                            # Ensure the merged bar catches up immediately on completion.
                            pb_gi = rep_to_group.get((sp, r))
                            if pb_gi is not None:
                                sc = step_counters.get((sp, r))
                                if sc is not None:
                                    cur = sc.value
                                    last = rep_last[(sp, r)]
                                    delta = cur - last
                                    if delta > 0:
                                        pbars[pb_gi].update(delta)
                                        rep_last[(sp, r)] = cur

                    # Aggregate any settings whose reps have all finished, and
                    # notify the caller as soon as each setting's entry is ready.
                    for sp in range(len(pending_settings)):
                        if sp in aggregated_settings:
                            continue
                        if not all((sp, r) in rep_results for r in range(n_repetitions)):
                            continue
                        global_idx_done = pending_settings[sp][0]
                        returns_list = [
                            np.asarray(rep_results[(sp, r)][0], dtype=np.float32)
                            for r in range(n_repetitions)
                        ]
                        ts = np.asarray(rep_results[(sp, 0)][1], dtype=np.int32)
                        all_ret = np.array(returns_list)
                        lc_mean = np.mean(all_ret, axis=0)
                        lc_std = (
                            np.std(all_ret, axis=0, ddof=1)
                            if len(returns_list) > 1 else np.zeros_like(lc_mean)
                        )
                        setting_results[global_idx_done] = (lc_mean, lc_std, ts, all_ret)
                        aggregated_settings.add(sp)
                        if on_setting_complete is not None:
                            try:
                                on_setting_complete(global_idx_done)
                            except Exception as cb_exc:
                                print(f"[on_setting_complete] callback raised: {cb_exc}")

                    time.sleep(0.25)
                    if poll_callback is not None:
                        try:
                            poll_callback()
                        except Exception as cb_exc:
                            print(f"[poll_callback] raised: {cb_exc}")
            finally:
                for pb in pbars.values():
                    pb.close()
                print()


# ── DQN repetitions via assignment2_repo ──────────────────────────────────────────────

def _run_dqn_repetitions(job_kwargs: dict[str, Any]):
    """Run n_repetitions of DQN via assignment2_repo and return (mean, std, timesteps).
    Uses ProcessPoolExecutor for parallel trials, matching assignment2_repo pattern."""
    from assignment2_repo.DQN import run_dqn_trial_returns

    kw: dict[str, Any] = dict(job_kwargs)
    n_repetitions = int(kw.pop("n_repetitions"))
    base_seed = int(kw.pop("base_seed", 42))
    n_timesteps = int(kw["n_timesteps"])
    eval_interval = int(kw.pop("eval_interval"))
    exploration_method = kw.pop("exploration_method")

    # Map field names to assignment2_repo DQN API
    trial_common: dict[str, Any] = dict(
        n_env_steps=n_timesteps,
        max_episode_length=int(kw.pop("max_episode_length")),
        learning_rate=float(kw.pop("learning_rate")),
        nn_hidden_layer_widths=np.asarray(kw.pop("nn_hidden_layer_widths"), dtype=np.int32),
        discount_factor=float(kw.pop("gamma")),
        target_network_step=int(kw.pop("TN_step")),
        target_network_active=bool(kw.pop("target_network_active")),
        n_returns_interval=eval_interval,
        exploration_method="epsilon_greedy" if exploration_method == "egreedy" else exploration_method,
        epsilon_start=kw.pop("epsilon_start", None),
        epsilon_end=kw.pop("epsilon_end", None),
        epsilon_decay=float(kw.pop("epsilon_decay", 1.0)),
        epsilon_decay_interval=int(kw.pop("epsilon_decay_interval", 1)),
        softmax_temp=kw.pop("softmax_temp", None),
        er_active=bool(kw.pop("er_active", False)),
        er_replay_buffer_size=int(kw.pop("er_replay_buffer_size", 10000)),
        er_batch_size=int(kw.pop("er_batch_size", 64)),
        er_min_replay_size=int(kw.pop("er_min_replay_size", 100)),
        er_sample_train_frequency=int(kw.pop("er_sample_train_frequency", 1)),
        er_replay_ratio=float(kw.pop("er_replay_ratio", 1.0)),
        enable_progress_bar=True,
        emit_trial_header=False,
    )
    # Remove n_timesteps from remaining kw (already consumed)
    kw.pop("n_timesteps", None)

    rep_outputs: list[tuple[np.ndarray, np.ndarray] | None] = [None for _ in range(n_repetitions)]
    parallel_workers = max(1, min(n_repetitions, os.cpu_count() or 1))

    if parallel_workers > 1 and n_repetitions > 1:
        with ProcessPoolExecutor(max_workers=parallel_workers) as executor:
            future_to_rep = {}
            for rep in range(n_repetitions):
                run_seed = base_seed + rep
                future = executor.submit(
                    run_dqn_trial_returns,
                    seed=run_seed,
                    trial_run_index=rep + 1,
                    total_trial_runs=n_repetitions,
                    progress_bar_position=rep,
                    progress_bar_desc=f"DQN Trial {rep + 1}/{n_repetitions}",
                    **trial_common,
                )
                future_to_rep[future] = rep
            for future in as_completed(future_to_rep):
                rep = future_to_rep[future]
                rep_outputs[rep] = future.result()
    else:
        for rep in range(n_repetitions):
            run_seed = base_seed + rep
            rep_outputs[rep] = run_dqn_trial_returns(
                seed=run_seed,
                trial_run_index=rep + 1,
                total_trial_runs=n_repetitions,
                progress_bar_position=rep,
                progress_bar_desc=f"DQN Trial {rep + 1}/{n_repetitions}",
                **trial_common,
            )

    completed_rep_outputs: list[tuple[np.ndarray, np.ndarray]] = [result for result in rep_outputs if result is not None]

    returns_list = []
    timesteps = None
    for returns, rep_ts in completed_rep_outputs:
        returns_list.append(np.asarray(returns, dtype=np.float32))
        if timesteps is None:
            timesteps = np.asarray(rep_ts, dtype=np.int32)

    all_returns = np.array(returns_list)
    lc_mean = np.mean(all_returns, axis=0)
    lc_std = (np.std(all_returns, axis=0, ddof=1)
              if all_returns.shape[0] > 1 else np.zeros_like(lc_mean))
    return lc_mean, lc_std, timesteps


# ── Workbook utilities (moved from excel_workbook_utils.py) ───────────────────

SAMPLE_TEMPLATE_PATH = os.path.join("data sheets", "Sample format", "1399.03.12.xlsx")
SETTING_SHEET_PREFIX = "Setting_"

TITLE_ROW_HEIGHT = 34.25
HEADER_ROW_HEIGHT = 52.5
DATA_ROW_HEIGHT = 23.4
DEFAULT_ROW_HEIGHT = 14.5
DEFAULT_COLUMN_WIDTH = 8.90625
SHEET_ZOOM_SCALE = 85


def build_algorithm_filename(algo_name: str) -> str:
    """Return the workbook filename stem for an algorithm."""
    return str(algo_name).upper()


def _normalize_for_signature(value: Any) -> Any:
    """Normalize values so workbook values and config values can be compared reliably."""
    if isinstance(value, np.ndarray):
        return tuple(_normalize_for_signature(v) for v in value.tolist())
    if isinstance(value, (list, tuple)):
        return tuple(_normalize_for_signature(v) for v in value)
    if isinstance(value, set):
        return tuple(sorted(_normalize_for_signature(v) for v in value))
    if isinstance(value, np.bool_):
        value = bool(value)
    elif isinstance(value, (np.integer, np.floating)):
        value = value.item()
    return _value_to_text(value)


def _value_to_text(value: Any) -> str:
    """Convert a value to a compact display string suitable for Excel text cells."""
    if value is None:
        return ""
    if isinstance(value, np.ndarray):
        return str(value.tolist())
    if isinstance(value, (list, tuple)):
        return str(list(value))
    if isinstance(value, set):
        return str(sorted(value))
    if isinstance(value, (np.integer, np.floating)):
        return _fmt(value.item())
    if isinstance(value, bool):
        return "True" if value else "False"
    return str(value)


def _job_signature(job_hyperparams: dict[str, Any]) -> tuple[tuple[str, Any], ...]:
    """Create a hashable signature from a job hyperparameter dictionary."""
    return tuple(sorted((str(key), _normalize_for_signature(value)) for key, value in job_hyperparams.items()))


def _sheet_signature(sheet_hyperparams: dict[str, Any]) -> tuple[tuple[str, Any], ...]:
    """Create a hashable signature from parsed workbook hyperparameters."""
    return tuple(sorted((str(key), _normalize_for_signature(value)) for key, value in sheet_hyperparams.items()))


def _entry_matches_job(entry_hyperparams: dict[str, Any], job_hyperparams: dict[str, Any]) -> bool:
    """Lenient de-dup: an existing entry matches a job if every key the entry
    has agrees with the job's corresponding value. Allows entries written before
    per-sheet global_config storage (a strict subset of today's keys) to still
    de-dup against jobs that now carry extra global_config keys."""
    for key, entry_val in entry_hyperparams.items():
        if key not in job_hyperparams:
            return False
        if _normalize_for_signature(entry_val) != _normalize_for_signature(job_hyperparams[key]):
            return False
    return True


@lru_cache(maxsize=1)
def _load_template_sheet():
    """Load the sample formatting sheet once and reuse it for all exports."""
    if not os.path.isfile(SAMPLE_TEMPLATE_PATH):
        return None
    workbook = load_workbook(SAMPLE_TEMPLATE_PATH)
    return workbook[workbook.sheetnames[0]]


def _copy_style(source_cell, target_cell) -> None:
    """Copy all visible style attributes from one cell to another."""
    target_cell._style = copy(source_cell._style)
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.protection = copy(source_cell.protection)
        target_cell.number_format = source_cell.number_format


def _apply_excel_sheet_shell(worksheet, sheet_title: str, n_columns: int) -> None:
    """Apply sheet-level settings that match the sample workbook."""
    worksheet.sheet_view.zoomScale = SHEET_ZOOM_SCALE
    worksheet.sheet_format.defaultRowHeight = DEFAULT_ROW_HEIGHT
    worksheet.sheet_format.defaultColWidth = DEFAULT_COLUMN_WIDTH
    worksheet.freeze_panes = None
    worksheet.auto_filter.ref = None
    worksheet.row_dimensions[1].height = TITLE_ROW_HEIGHT
    worksheet.row_dimensions[2].height = HEADER_ROW_HEIGHT
    for row_index in range(3, worksheet.max_row + 1):
        worksheet.row_dimensions[row_index].height = DATA_ROW_HEIGHT


def _format_excel(
    worksheet,
    *,
    algo_name: str,
    hyperparams: dict[str, Any],
    headers: list[str],
    rows: list[list[Any]],
) -> None:
    """Format a worksheet to match the attached sample workbook style."""
    template_sheet = _load_template_sheet()
    n_columns = max(1, len(headers))

    _apply_excel_sheet_shell(worksheet, algo_name, n_columns)

    title_text = f"{str(algo_name).upper()}: " + ", ".join(
        f"{key}={_value_to_text(value)}" for key, value in hyperparams.items()
    )
    title_end = get_column_letter(n_columns)
    worksheet.merge_cells(f"A1:{title_end}1")
    worksheet["A1"] = title_text

    if template_sheet is not None:
        _copy_style(template_sheet["A1"], worksheet["A1"])
    else:
        worksheet["A1"].alignment = worksheet["A1"].alignment.copy(horizontal="center", vertical="center")
        worksheet["A1"].font = worksheet["A1"].font.copy(bold=True)

    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(2, column_index, header)
        if template_sheet is not None:
            template_col = min(column_index, template_sheet.max_column)
            _copy_style(template_sheet.cell(2, template_col), cell)
        else:
            cell.alignment = cell.alignment.copy(horizontal="center", vertical="center", wrap_text=True)
            cell.font = cell.font.copy(bold=True)

    for row_index, row_values in enumerate(rows, start=3):
        worksheet.row_dimensions[row_index].height = DATA_ROW_HEIGHT
        for column_index, value in enumerate(row_values, start=1):
            cell = worksheet.cell(row_index, column_index, value)
            if template_sheet is not None:
                template_col = min(column_index, template_sheet.max_column)
                _copy_style(template_sheet.cell(3, template_col), cell)
            else:
                cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
                cell.font = cell.font.copy(bold=True)

    if rows:
        first_data_row = rows[0]
    else:
        first_data_row = ["" for _ in headers]

    for column_index, header in enumerate(headers, start=1):
        first_value = first_data_row[column_index - 1] if column_index - 1 < len(first_data_row) else ""
        width = max(_excel_cell_display_width(header), _excel_cell_display_width(first_value)) + 1
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    worksheet.sheet_view.tabSelected = True


def _write_raw_excel_sheet(worksheet, headers: list[str], rows: list[list[Any]]) -> None:
    """Write a worksheet without merged cells or style formatting."""
    for column_index, header in enumerate(headers, start=1):
        worksheet.cell(1, column_index, header)

    for row_index, row_values in enumerate(rows, start=2):
        for column_index, value in enumerate(row_values, start=1):
            worksheet.cell(row_index, column_index, value)

    first_data_row = rows[0] if rows else []
    _set_raw_excel_column_widths(worksheet, headers, first_data_row)


def _next_setting_sheet_name(workbook) -> str:
    """Return the next available Setting_### sheet name."""
    highest_index = 0
    for sheet_name in workbook.sheetnames:
        if not sheet_name.startswith(SETTING_SHEET_PREFIX):
            continue
        suffix = sheet_name[len(SETTING_SHEET_PREFIX):]
        try:
            highest_index = max(highest_index, int(suffix))
        except ValueError:
            continue
    return f"{SETTING_SHEET_PREFIX}{highest_index + 1:03d}"


def _rows_from_result(
    result: tuple[np.ndarray, np.ndarray, np.ndarray]
    | tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray]
) -> list[list[Any]]:
    """Convert a (mean,std,timesteps[,raw_returns]) tuple into row-wise Excel rows.

    If 'raw_returns' is provided it is expected to have shape:
        (n_repetitions, n_points)

    Then each row becomes:
        [timestep, mean, std, rep_1, ..., rep_n]
    """
    learning_curve, learning_curve_std, timesteps = result[:3]
    raw_returns = result[3] if len(result) == 4 else None

    timesteps_arr = np.asarray(timesteps)
    learning_curve_arr = np.asarray(learning_curve)
    learning_curve_std_arr = np.asarray(learning_curve_std)

    if raw_returns is None:
        rows = []
        for timestep, mean_value, std_value in zip(
            timesteps_arr, learning_curve_arr, learning_curve_std_arr
        ):
            rows.append([timestep, mean_value, std_value])
        return rows

    raw_returns_arr = np.asarray(raw_returns, dtype=np.float32)

    n_points = min(
        len(timesteps_arr),
        len(learning_curve_arr),
        len(learning_curve_std_arr),
        raw_returns_arr.shape[1],
    )

    rows = []
    for idx in range(n_points):
        rep_values = raw_returns_arr[:, idx].tolist()
        rows.append([
            timesteps_arr[idx],
            learning_curve_arr[idx],
            learning_curve_std_arr[idx],
            *rep_values,
        ])
    return rows


def _build_headers(job_hyperparams: dict[str, Any]) -> list[str]:
    """Build the Excel column headers for a saved setting."""
    try:
        n_repetitions = int(job_hyperparams.get("n_repetitions", 0))
    except (TypeError, ValueError):
        n_repetitions = 0

    rep_headers = [f"rep_{i + 1}" for i in range(max(0, n_repetitions))]

    return [
        "timestep",
        "learning_curve_mean",
        "learning_curve_std",
        *rep_headers,
        *[str(key) for key in job_hyperparams.keys()],
        "curve_label",
    ]


def _build_rows(
    job_hyperparams: dict[str, Any],
    curve_label: str,
    result: tuple[np.ndarray, np.ndarray, np.ndarray]
    | tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray],
) -> list[list[Any]]:
    """Build the Excel rows for a single setting sheet.

    Writes:
    - timestep / mean / std for every row
    - rep_1..rep_n for every row (when 'raw_returns' is available)
    - hyperparameters + curve_label ONLY on the first timestep row
    """
    base_rows = _rows_from_result(result)

    try:
        n_repetitions = int(job_hyperparams.get("n_repetitions", 0))
    except (TypeError, ValueError):
        n_repetitions = 0
    n_repetitions = max(0, n_repetitions)

    hyperparam_headers = [str(k) for k in job_hyperparams.keys()]
    hyperparam_values = [_value_to_text(job_hyperparams[h]) for h in hyperparam_headers]

    rows: list[list[Any]] = []
    hyperparam_count = len(hyperparam_values)

    for row_index, base_row in enumerate(base_rows):
        # base_row is either:
        #  - [timestep, mean, std]
        #  - [timestep, mean, std, rep_1, ..., rep_n]
        timestep = base_row[0]
        mean_value = base_row[1]
        std_value = base_row[2]
        rep_values = base_row[3:] if len(base_row) > 3 else []

        if n_repetitions > 0:
            if len(rep_values) < n_repetitions:
                rep_values = [None] * (n_repetitions - len(rep_values)) + rep_values
                rep_values = rep_values[-n_repetitions:]
            elif len(rep_values) > n_repetitions:
                rep_values = rep_values[:n_repetitions]
        else:
            rep_values = []

        if row_index == 0:
            rows.append([timestep, mean_value, std_value, *rep_values, *hyperparam_values, curve_label])
        else:
            rows.append([
                timestep,
                mean_value,
                std_value,
                *rep_values,
                *([None] * hyperparam_count),
                "",
            ])

    return rows


def _parse_sheet_entry(worksheet, *, formatted_sheets: bool = False) -> dict[str, Any] | None:
    """Parse one worksheet into the in-memory result format."""
    header_row = 2 if formatted_sheets else 1
    data_start_row = 3 if formatted_sheets else 2

    headers: list[str] = []
    header_to_column: dict[str, int] = {}

    for column_index in range(1, worksheet.max_column + 1):
        header_value = worksheet.cell(header_row, column_index).value
        if header_value is None:
            continue
        header_text = str(header_value)
        if header_text in header_to_column:
            continue
        header_to_column[header_text] = column_index
        headers.append(header_text)

    required = {"timestep", "learning_curve_mean", "learning_curve_std"}
    if not required.issubset(set(headers)):
        return None

    column_data: dict[str, list[Any]] = {header: [] for header in headers}
    for row_index in range(data_start_row, worksheet.max_row + 1):
        non_empty = False
        for header in headers:
            column_index = header_to_column[header]
            value = worksheet.cell(row_index, column_index).value
            if value is not None:
                non_empty = True
            column_data[header].append(value)
        if not non_empty:
            continue

    try:
        timesteps = np.asarray(column_data["timestep"], dtype=np.int32)
        learning_curve = np.asarray(column_data["learning_curve_mean"], dtype=np.float32)
        learning_curve_std = np.asarray(column_data["learning_curve_std"], dtype=np.float32)
    except Exception:
        return None

    rep_headers = [h for h in headers if h.startswith("rep_")]
    rep_headers_sorted = sorted(
        rep_headers,
        key=lambda s: int(s.split("_", 1)[1]) if s.split("_", 1)[1].isdigit() else 10**9,
    )

    raw_returns = None
    if rep_headers_sorted:
        # raw_returns shape: (n_repetitions, n_points)
        rep_cols = []
        for h in rep_headers_sorted:
            col = column_data[h]
            rep_cols.append([np.nan if v is None else v for v in col])
        raw_returns = np.asarray(rep_cols, dtype=np.float32)

    hyperparams: dict[str, Any] = {}
    for header in headers:
        # rep_* must never be part of the hyperparameter signature
        if header.startswith("rep_"):
            continue
        if header in required or header == "curve_label":
            continue
        values = column_data[header]
        first_value = values[0] if values else None
        if first_value is not None:
            hyperparams[header] = _value_to_text(first_value)

    curve_values = column_data.get("curve_label", [])
    curve_label = next((str(value) for value in curve_values if value not in (None, "")), "")

    return {
        "learning_curve": learning_curve,
        "learning_curve_std": learning_curve_std,
        "timesteps": timesteps,
        "raw_returns": raw_returns,
        "curve_label": curve_label,
        "hyperparams": hyperparams,
    }


def load_algorithm_workbook(
    filepath: str,
    *,
    setting_jobs: list[dict[str, Any]] | None = None,
    formatted_sheets: bool = False,
) -> tuple[list[dict[str, Any] | None], dict[str, tuple[Any, Any]]]:
    """Load matching sheets from a workbook."""
    try:
        workbook = load_workbook(filepath)
    except Exception as exc:
        raise ValueError(f"Failed to read Excel file '{filepath}': {exc}") from exc

    parsed_entries: list[dict[str, Any] | None] = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        entry = _parse_sheet_entry(worksheet, formatted_sheets=formatted_sheets)
        if entry is None:
            entry = _parse_sheet_entry(worksheet, formatted_sheets=not formatted_sheets)
        if entry is not None:
            parsed_entries.append(entry)

    if setting_jobs is None:
        return parsed_entries, {}

    job_map: dict[tuple[tuple[str, Any], ...], int] = {}
    for index, job in enumerate(setting_jobs):
        job_map[_job_signature(job["hyperparams"])] = index

    aligned_entries: list[dict[str, Any] | None] = [None] * len(setting_jobs)
    for entry in parsed_entries:
        signature = _sheet_signature(entry["hyperparams"])
        job_index = job_map.get(signature)
        if job_index is not None and aligned_entries[job_index] is None:
            aligned_entries[job_index] = entry

    return aligned_entries, {}


def save_algorithm_workbook(
    dir_path: str,
    base_filename: str,
    algo_name: str,
    setting_jobs: list[dict[str, Any]],
    setting_results: list[tuple[np.ndarray, np.ndarray, np.ndarray] | None],
    format_sheets: bool = False,
    verbose: bool = True,
    *,
    global_config: dict[str, Any] | None = None,
    algo_config: dict[str, Any] | None = None,
) -> str:
    """Save or append settings into an algorithm workbook.

    Each setting sheet records both the swept algo hyperparameters and the
    filtered global_config values that were active when the curve was produced,
    so that future runs can validate each sheet independently of the others in
    the same workbook.
    """
    del algo_config  # accepted for caller compatibility; not used at save time.

    os.makedirs(dir_path, exist_ok=True)
    filepath = os.path.join(dir_path, f"{base_filename}.xlsx")

    global_filtered = _meta_filtered_items(global_config, GLOBAL_CONFIG_EXCLUSIONS) if global_config else {}

    existing_entries: list[dict[str, Any]] = []
    if os.path.isfile(filepath):
        try:
            existing_entries, _ = load_algorithm_workbook(filepath, formatted_sheets=format_sheets)
        except Exception:
            existing_entries = []

    # Back-compat: existing sheets predating per-sheet global_config storage are
    # migrated by assuming the current global_config values applied to them.
    for entry in existing_entries:
        entry_hp = entry["hyperparams"]
        for gc_key, gc_text in global_filtered.items():
            if gc_key not in entry_hp:
                entry_hp[gc_key] = gc_text

    all_entries = list(existing_entries)
    added_count = 0
    for job, result in zip(setting_jobs, setting_results):
        if result is None:
            continue

        job_hp = dict(job["hyperparams"])
        for gc_key, gc_text in global_filtered.items():
            job_hp.setdefault(gc_key, gc_text)

        if any(_entry_matches_job(entry["hyperparams"], job_hp) for entry in all_entries):
            continue

        learning_curve, learning_curve_std, timesteps = result[:3]
        raw_returns = result[3] if len(result) == 4 else None
        all_entries.append({
            "learning_curve": learning_curve,
            "learning_curve_std": learning_curve_std,
            "timesteps": timesteps,
            "raw_returns": raw_returns,
            "curve_label": job["curve_label"],
            "hyperparams": job_hp,
        })
        added_count += 1

    workbook = Workbook()
    default_sheet = workbook.active
    if default_sheet is not None:
        workbook.remove(default_sheet)

    if not all_entries:
        worksheet = workbook.create_sheet(title=_next_setting_sheet_name(workbook))
        if format_sheets:
            _format_excel(
                worksheet,
                algo_name=algo_name,
                hyperparams={},
                headers=["timestep", "learning_curve_mean", "learning_curve_std", "curve_label"],
                rows=[],
            )
        else:
            _write_raw_excel_sheet(
                worksheet,
                ["timestep", "learning_curve_mean", "learning_curve_std", "curve_label"],
                [],
            )
    else:
        for index, entry in enumerate(all_entries, start=1):
            worksheet = workbook.create_sheet(title=f"Setting_{index:03d}")
            headers = _build_headers(entry["hyperparams"])
            raw_returns = entry.get("raw_returns")
            if raw_returns is None:
                result_tuple = (
                    entry["learning_curve"],
                    entry["learning_curve_std"],
                    entry["timesteps"],
                )
            else:
                result_tuple = (
                    entry["learning_curve"],
                    entry["learning_curve_std"],
                    entry["timesteps"],
                    raw_returns,
                )

            rows = _build_rows(entry["hyperparams"], entry["curve_label"], result_tuple)
            if format_sheets:
                _format_excel(
                    worksheet,
                    algo_name=algo_name,
                    hyperparams=entry["hyperparams"],
                    headers=headers,
                    rows=rows,
                )
            else:
                _write_raw_excel_sheet(worksheet, headers, rows)

    workbook.save(filepath)
    if verbose:
        print(f"Saved {added_count} new setting(s) to {filepath}")
    return filepath, added_count


# ── Returns summary table ─────────────────────────────────────────────────────

def _pooled_stats_from_curves(
    learning_curve: np.ndarray,
    learning_curve_std: np.ndarray,
    n_repetitions: int,
) -> tuple[float, float, int]:
    """Reconstruct (mean, population-std, n_samples) from per-step mean/sample-std.

    Used when per-rep raw returns are not available on disk. Treats
    'learning_curve_std' as a sample std with ddof=1 across reps at each step.
    """
    lc = np.asarray(learning_curve, dtype=np.float64).reshape(-1)
    ls = np.asarray(learning_curve_std, dtype=np.float64).reshape(-1)
    if lc.size == 0:
        return float("nan"), float("nan"), 0
    n = max(int(n_repetitions), 1)
    var_pop_per_step = ((n - 1) / n) * (ls ** 2) if n > 1 else np.zeros_like(lc)
    e_x2 = float(np.mean(lc ** 2 + var_pop_per_step))
    m = float(np.mean(lc))
    var_total = max(e_x2 - m * m, 0.0)
    return m, float(np.sqrt(var_total)), int(lc.size * n)


def _format_number(value: float, *, decimals: int = 2) -> str:
    if value is None or not np.isfinite(value):
        return "n/a"
    return f"{value:,.{decimals}f}"


def _render_aligned_table(headers: list[str], rows: list[list[str]]) -> str:
    """Render a plain-text table with aligned columns (right-align numeric cols)."""
    if not rows:
        return ""
    widths = [len(h) for h in headers]
    for row in rows:
        for i, cell in enumerate(row):
            widths[i] = max(widths[i], len(cell))

    def _fmt_row(cells: list[str]) -> str:
        out = []
        for i, cell in enumerate(cells):
            if i <= 1:
                out.append(cell.ljust(widths[i]))
            else:
                out.append(cell.rjust(widths[i]))
        return "  ".join(out)

    sep = "  ".join("-" * w for w in widths)
    lines = [_fmt_row(headers), sep]
    for row in rows:
        lines.append(_fmt_row(row))
    return "\n".join(lines)


def _non_excluded_hp_items(job_hp: dict[str, Any]) -> list[tuple[str, str]]:
    """Return sorted (key, stable-text) items from job hyperparams with structural
    exclusions removed. Independent of 'legend_parameters' show flags.
    """
    excluded = GLOBAL_CONFIG_EXCLUSIONS | ALGO_CONFIG_EXCLUSIONS
    items: list[tuple[str, str]] = []
    for k, v in job_hp.items():
        key = str(k)
        if key in excluded:
            continue
        items.append((key, _meta_value_text(v)))
    items.sort(key=lambda kv: kv[0])
    return items


def _format_setting_label(algo_upper: str, hp_items: list[tuple[str, str]]) -> str:
    parts = [algo_upper]
    for k, v in hp_items:
        parts.append(f"{k}={v}")
    return ", ".join(parts)


def build_returns_summary_table(
    *,
    algo_jobs: dict[str, list[dict[str, Any]]],
    setting_results: list,
    algo_job_offsets: dict[str, int],
    n_repetitions: int,
    last_fraction: float = 0.1,
    output_dir: str = ".",
    csv_filename: str = "results_summary.csv",
    md_filename: str = "results_summary.md",
    print_to_stdout: bool = True,
) -> dict[str, Any]:
    """Aggregate returns per (algorithm, non-excluded-HP) row and emit a table.

    Settings sharing identical non-excluded hyperparameters (after removing only
    GLOBAL_CONFIG_EXCLUSIONS / ALGO_CONFIG_EXCLUSIONS) collapse into one row.
    Group identity and the "Setting" label are derived from every non-excluded
    hyperparameter in 'job["hyperparams"]', independent of 'legend_parameters'.
    All repetitions and all eval points are pooled into the overall mean/std;
    the last 'last_fraction' of eval points produces the 'last N%' columns.
    """
    last_fraction = float(last_fraction)
    last_fraction = min(max(last_fraction, 0.0), 1.0)

    grouped: dict[tuple[str, tuple[tuple[str, str], ...]], dict[str, Any]] = {}
    order: list[tuple[str, tuple[tuple[str, str], ...]]] = []

    for algo_upper, jobs in algo_jobs.items():
        offset = algo_job_offsets[algo_upper]
        for i, job in enumerate(jobs):
            entry = setting_results[offset + i]
            if entry is None:
                continue
            hp_items = _non_excluded_hp_items(job.get("hyperparams", {}))
            setting_label = _format_setting_label(algo_upper, hp_items)
            key = (algo_upper, tuple(hp_items))
            if key not in grouped:
                grouped[key] = {
                    "algo": algo_upper,
                    "label": setting_label,
                    "all_values": [],
                    "last_values": [],
                    "n_settings": 0,
                    "missing_raw": 0,
                }
                order.append(key)

            lc = np.asarray(entry[0], dtype=np.float64).reshape(-1)
            ls = np.asarray(entry[1], dtype=np.float64).reshape(-1)
            raw = entry[3] if len(entry) >= 4 else None

            grouped[key]["n_settings"] += 1

            if raw is not None and getattr(raw, "size", 0) > 0:
                raw_arr = np.asarray(raw, dtype=np.float64)
                if raw_arr.ndim == 1:
                    raw_arr = raw_arr.reshape(1, -1)
                grouped[key]["all_values"].append(raw_arr.reshape(-1))
                n_pts = raw_arr.shape[1]
                n_last = max(1, int(np.ceil(last_fraction * n_pts))) if last_fraction > 0 else 0
                if n_last > 0:
                    grouped[key]["last_values"].append(raw_arr[:, -n_last:].reshape(-1))
            else:
                grouped[key]["missing_raw"] += 1
                # Fallback: synthesize from (lc, lc_std) treated as per-step rep mean/std.
                n = max(int(n_repetitions), 1)
                grouped[key].setdefault("fallback_curves_all", []).append((lc, ls, n))
                if last_fraction > 0 and lc.size > 0:
                    n_last = max(1, int(np.ceil(last_fraction * lc.size)))
                    grouped[key].setdefault("fallback_curves_last", []).append(
                        (lc[-n_last:], ls[-n_last:], n)
                    )

    def _aggregate(values_list, fallback_list):
        if values_list:
            flat = np.concatenate([v for v in values_list if v.size > 0]) if values_list else np.array([])
        else:
            flat = np.array([])
        if fallback_list:
            # Reconstruct first/second moments per fallback chunk and combine.
            total_n = 0
            sum_x = 0.0
            sum_x2 = 0.0
            if flat.size > 0:
                total_n += flat.size
                sum_x += float(np.sum(flat))
                sum_x2 += float(np.sum(flat * flat))
            for lc, ls, n in fallback_list:
                if lc.size == 0:
                    continue
                var_pop_per_step = ((n - 1) / n) * (ls ** 2) if n > 1 else np.zeros_like(lc)
                chunk_n = lc.size * n
                chunk_sum = float(np.sum(lc)) * n
                chunk_sum_x2 = float(np.sum(lc ** 2 + var_pop_per_step)) * n
                total_n += chunk_n
                sum_x += chunk_sum
                sum_x2 += chunk_sum_x2
            if total_n == 0:
                return float("nan"), float("nan"), 0
            mean = sum_x / total_n
            var = max(sum_x2 / total_n - mean * mean, 0.0)
            return mean, float(np.sqrt(var)), total_n
        if flat.size == 0:
            return float("nan"), float("nan"), 0
        return float(np.mean(flat)), float(np.std(flat, ddof=0)), int(flat.size)

    rows_data: list[dict[str, Any]] = []
    for key in order:
        info = grouped[key]
        mean_all, std_all, n_all = _aggregate(
            info["all_values"], info.get("fallback_curves_all", [])
        )
        mean_last, std_last, n_last = _aggregate(
            info["last_values"], info.get("fallback_curves_last", [])
        )
        rows_data.append({
            "algorithm": info["algo"],
            "setting": info["label"],
            "mean_all": mean_all,
            "std_all": std_all,
            "n_all": n_all,
            "mean_last": mean_last,
            "std_last": std_last,
            "n_last": n_last,
            "n_settings": info["n_settings"],
            "missing_raw": info["missing_raw"],
        })

    last_pct = int(round(last_fraction * 100))
    headers = [
        "Algorithm",
        "Setting",
        "Mean (all)",
        "Std (all)",
        f"Mean (last {last_pct}%)",
        f"Std (last {last_pct}%)",
        "N (all)",
        "N (last)",
    ]
    text_rows: list[list[str]] = []
    csv_rows: list[list[str]] = []
    md_rows: list[list[str]] = []
    for row in rows_data:
        text_rows.append([
            row["algorithm"],
            row["setting"],
            _format_number(row["mean_all"]),
            _format_number(row["std_all"]),
            _format_number(row["mean_last"]),
            _format_number(row["std_last"]),
            f"{row['n_all']:,}",
            f"{row['n_last']:,}",
        ])
        csv_rows.append([
            row["algorithm"],
            row["setting"],
            f"{row['mean_all']:.6f}" if np.isfinite(row["mean_all"]) else "",
            f"{row['std_all']:.6f}" if np.isfinite(row["std_all"]) else "",
            f"{row['mean_last']:.6f}" if np.isfinite(row["mean_last"]) else "",
            f"{row['std_last']:.6f}" if np.isfinite(row["std_last"]) else "",
            str(row["n_all"]),
            str(row["n_last"]),
        ])
        md_rows.append(text_rows[-1])

    rendered_text = _render_aligned_table(headers, text_rows) if text_rows else "(no completed settings to summarize)"

    title_stdout = f"Returns summary - mean/std across all repetitions and eval points (n_repetitions={n_repetitions})"
    title_md = f"Returns summary - mean ± std across all repetitions and eval points (n_repetitions={n_repetitions})"

    # ── Console (ASCII-safe for Windows cp1252 terminals) ──
    if print_to_stdout:
        def _safe_print(s: str) -> None:
            try:
                print(s)
            except UnicodeEncodeError:
                print(s.encode("ascii", errors="replace").decode("ascii"))

        _safe_print("\n" + "=" * len(title_stdout))
        _safe_print(title_stdout)
        _safe_print("=" * len(title_stdout))
        _safe_print(rendered_text)
        _safe_print("")

    os.makedirs(output_dir, exist_ok=True)

    # ── Markdown ──
    md_path = os.path.join(output_dir, md_filename)
    md_lines = [f"# {title_md}", ""]
    if md_rows:
        md_lines.append("| " + " | ".join(headers) + " |")
        align = ["---", "---", "---:", "---:", "---:", "---:", "---:", "---:"]
        md_lines.append("| " + " | ".join(align) + " |")
        for row in md_rows:
            md_lines.append("| " + " | ".join(row) + " |")
    else:
        md_lines.append("_(no completed settings to summarize)_")
    with open(md_path, "w", encoding="utf-8") as f:
        f.write("\n".join(md_lines) + "\n")

    # ── CSV ──
    csv_path = os.path.join(output_dir, csv_filename)
    import csv as _csv
    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        writer = _csv.writer(f)
        writer.writerow(headers)
        for row in csv_rows:
            writer.writerow(row)

    if print_to_stdout:
        print(f"Saved returns summary to: {md_path}")
        print(f"Saved returns summary to: {csv_path}")
        print()

    return {
        "rows": rows_data,
        "headers": headers,
        "markdown_path": md_path,
        "csv_path": csv_path,
        "text": rendered_text,
    }
